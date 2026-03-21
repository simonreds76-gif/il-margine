#!/usr/bin/env python3
from __future__ import annotations

import csv
from collections import OrderedDict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
EXPORT_DIR = ROOT / "data" / "exports"
DEFAULT_OUTPUT = EXPORT_DIR / "betting-archive.xlsx"


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def infer_goalscorer_league(path: Path) -> str:
    stem = path.stem.lower()
    if stem.startswith("epl-"):
        return "epl"
    if stem.startswith("bundesliga-"):
        return "bundesliga"
    if stem.startswith("la-liga-"):
        return "la-liga"
    if stem.startswith("ligue-1-"):
        return "ligue-1"
    if stem.startswith("serie-a-"):
        return "serie-a"
    return "serie-a"


def dedupe_rows(rows: Iterable[dict[str, str]], keys: list[str]) -> list[dict[str, str]]:
    deduped: OrderedDict[str, dict[str, str]] = OrderedDict()
    for row in rows:
        identity = "|".join((row.get(key) or "").strip() for key in keys)
        deduped[identity] = row
    return list(deduped.values())


def autosize_sheet(worksheet) -> None:
    widths: dict[int, int] = {}
    for row in worksheet.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(value) + 2), 42)
    for column_idx, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column_idx)].width = width


def write_sheet(workbook: Workbook, title: str, rows: list[dict[str, str]]) -> None:
    if title in workbook.sheetnames:
        worksheet = workbook[title]
        if worksheet.max_row > 0:
            worksheet.delete_rows(1, worksheet.max_row)
    else:
        worksheet = workbook.create_sheet(title)

    if not rows:
        worksheet.append(["no_rows"])
        autosize_sheet(worksheet)
        return

    headers = list(rows[0].keys())
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize_sheet(worksheet)


def build_goalscorer_shadow_rows() -> list[dict[str, str]]:
    goalscorer_files = sorted((ROOT / "data" / "goalscorer").glob("*shadow-signals.csv"))
    combined: list[dict[str, str]] = []
    for path in goalscorer_files:
        for row in read_csv_rows(path):
            enriched = dict(row)
            enriched["archive_league"] = infer_goalscorer_league(path)
            enriched["archive_source_file"] = path.name
            combined.append(enriched)

    ordered = dedupe_rows(
        combined,
        ["date", "kickoff", "match", "player", "signal_type", "archive_league"],
    )
    ordered.sort(
        key=lambda row: (
            row.get("date", ""),
            row.get("kickoff", ""),
            row.get("archive_league", ""),
            row.get("player", ""),
        ),
        reverse=True,
    )
    return ordered


def build_backup_rows(pattern: str) -> list[dict[str, str]]:
    backup_files = sorted((ROOT / "data" / "backtest").glob(pattern))
    combined: list[dict[str, str]] = []
    for path in backup_files:
        stem = path.stem
        stamp = stem.split(".bak-")[-1] if ".bak-" in stem else ""
        for row in read_csv_rows(path):
            enriched = dict(row)
            enriched["archive_source_file"] = path.name
            enriched["archive_snapshot"] = stamp
            combined.append(enriched)

    combined.sort(
        key=lambda row: (
            row.get("archive_snapshot", ""),
            row.get("date", ""),
            row.get("time_utc", ""),
            row.get("player1", ""),
            row.get("player", ""),
        ),
        reverse=True,
    )
    return combined


def main() -> None:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    readme_sheet = workbook.active
    readme_sheet.title = "README"

    fair_odds_sources = [
        ("fair_odds_base", ROOT / "data" / "backtest" / "strict-signals.csv"),
        ("fair_odds_overlay", ROOT / "data" / "backtest" / "strict-signals-overlay-compare.csv"),
        ("fair_odds_volume200", ROOT / "data" / "backtest" / "strict-signals-volume200.csv"),
        ("fair_odds_vol200_int", ROOT / "data" / "backtest" / "strict-signals-volume200-internal.csv"),
        ("fair_odds_internal5", ROOT / "data" / "backtest" / "strict-signals-internal-5pct.csv"),
        ("fair_odds_volume275", ROOT / "data" / "backtest" / "strict-signals-volume275.csv"),
    ]
    backup_sources = [
        ("bak_base", "strict-signals.bak-*.csv"),
        ("bak_overlay", "strict-signals-overlay-compare.bak-*.csv"),
        ("bak_volume200", "strict-signals-volume200.bak-*.csv"),
        ("bak_internal5", "strict-signals-internal-5pct.bak-*.csv"),
    ]

    readme_rows: list[dict[str, str]] = []
    for sheet_name, path in fair_odds_sources:
        rows = read_csv_rows(path)
        write_sheet(workbook, sheet_name, rows)
        readme_rows.append(
            {
                "sheet": sheet_name,
                "source_file": str(path.relative_to(ROOT)),
                "rows": str(len(rows)),
                "notes": "Current canonical fair-odds archive file",
            }
        )

    goalscorer_shadow_rows = build_goalscorer_shadow_rows()
    write_sheet(workbook, "goalscorer_shadow", goalscorer_shadow_rows)
    readme_rows.append(
        {
            "sheet": "goalscorer_shadow",
            "source_file": "data/goalscorer/*shadow-signals.csv",
            "rows": str(len(goalscorer_shadow_rows)),
            "notes": "Combined and deduped goalscorer shadow archive across league files",
        }
    )

    for sheet_name, pattern in backup_sources:
        rows = build_backup_rows(pattern)
        write_sheet(workbook, sheet_name, rows)
        readme_rows.append(
            {
                "sheet": sheet_name,
                "source_file": f"data/backtest/{pattern}",
                "rows": str(len(rows)),
                "notes": "Historical backup snapshots for this fair-odds archive family",
            }
        )

    write_sheet(workbook, "README", readme_rows)

    workbook.save(DEFAULT_OUTPUT)
    print(f"Saved workbook: {DEFAULT_OUTPUT}")


if __name__ == "__main__":
    main()
