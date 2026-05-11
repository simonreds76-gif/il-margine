#!/usr/bin/env python3
"""
Fetch/update ATP season file from tennis-data.co.uk and validate Pinnacle coverage.

Purpose:
- Keep local `data/backtest/atp-YYYY.xlsx` current for historical backfill.
- Validate whether Pinnacle columns (PSW/PSL) are actually present beyond the
  latest settled strict sample.

This is a backfill/verification path, not the primary live CLV source.

Usage:
  python scripts/fetch-tennis-data-atp.py
  python scripts/fetch-tennis-data-atp.py --year 2026
  python scripts/fetch-tennis-data-atp.py --dry-run
"""

from __future__ import annotations

import argparse
import re
import zipfile
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import requests
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data" / "backtest"
ALDATA_URL = "https://www.tennis-data.co.uk/alldata.php"
REQUEST_TIMEOUT = 30


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fetch/update ATP season file from tennis-data.co.uk")
    parser.add_argument("--year", type=int, default=date.today().year)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--report", default=str(DATA_DIR / "tennis-data-atp-refresh.txt"))
    return parser.parse_args()


def find_candidate_urls(year: int) -> list[str]:
    candidates: list[str] = []
    index_error: Exception | None = None
    fallbacks = [
        # tennis-data HTTPS is flaky on some Windows TLS stacks; direct HTTP is
        # intentionally first because the workbook endpoint still serves 200.
        f"http://www.tennis-data.co.uk/{year}/{year}.xlsx",
        f"http://www.tennis-data.co.uk/{year}/{year}.zip",
        f"https://www.tennis-data.co.uk/{year}/{year}.xlsx",
        f"https://www.tennis-data.co.uk/{year}/{year}.zip",
    ]
    try:
        resp = requests.get(ALDATA_URL, timeout=REQUEST_TIMEOUT, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        html = resp.text
        hrefs = re.findall(r'href=["\']([^"\']+)["\']', html, flags=re.I)
        for href in hrefs:
            full = urljoin(ALDATA_URL, href)
            if str(year) not in full:
                continue
            if not re.search(rf"/{year}/.*\.(zip|xlsx|xls)$", full, flags=re.I):
                continue
            candidates.append(full)
    except Exception as exc:
        index_error = exc

    out: list[str] = []
    seen = set()
    for url in [*fallbacks, *candidates]:
        if url in seen:
            continue
        seen.add(url)
        out.append(url)
    if not out and index_error:
        raise RuntimeError(f"Could not read tennis-data index or build fallbacks: {index_error}") from index_error
    return out


def choose_payload(resp: requests.Response, year: int) -> tuple[bytes, str, str]:
    url = resp.url
    data = resp.content
    lower_url = url.lower()

    if lower_url.endswith(".xlsx"):
        return data, ".xlsx", url
    if lower_url.endswith(".zip"):
        with zipfile.ZipFile(BytesIO(data)) as zf:
            names = zf.namelist()
            xlsx_members = [n for n in names if str(year) in n and n.lower().endswith(".xlsx")]
            if xlsx_members:
                member = sorted(xlsx_members)[0]
                return zf.read(member), ".xlsx", f"{url}::{member}"
            xls_members = [n for n in names if str(year) in n and n.lower().endswith(".xls")]
            if xls_members:
                member = sorted(xls_members)[0]
                return zf.read(member), ".xls", f"{url}::{member}"
        raise RuntimeError(f"No season workbook found inside zip: {url}")
    if lower_url.endswith(".xls"):
        return data, ".xls", url
    raise RuntimeError(f"Unsupported tennis-data payload type: {url}")


def download_season_file(year: int) -> tuple[bytes, str, str]:
    errors: list[str] = []
    for url in find_candidate_urls(year):
        try:
            resp = requests.get(url, timeout=REQUEST_TIMEOUT, headers={"User-Agent": "Mozilla/5.0"})
            if not resp.ok:
                errors.append(f"{url} -> HTTP {resp.status_code}")
                continue
            return choose_payload(resp, year)
        except Exception as exc:
            errors.append(f"{url} -> {exc}")
    raise RuntimeError("No usable tennis-data season file found.\n" + "\n".join(errors))


def inspect_xlsx(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(h) if h is not None else "" for h in next(rows)]
    idx = {k: i for i, k in enumerate(header)}
    required = {"Date", "Comment", "PSW", "PSL"}
    missing = sorted(required - set(idx))
    if missing:
        raise RuntimeError(f"{path.name}: missing required columns {missing}")

    total_rows = 0
    completed_rows = 0
    ps_rows = 0
    all_dates: list[date] = []
    ps_dates: list[date] = []

    for row in rows:
        total_rows += 1
        dt = row[idx["Date"]]
        if isinstance(dt, datetime):
            d = dt.date()
        elif isinstance(dt, date):
            d = dt
        else:
            continue
        all_dates.append(d)
        comment = str(row[idx["Comment"]] or "").strip()
        if comment == "Completed":
            completed_rows += 1
            psw = row[idx["PSW"]]
            psl = row[idx["PSL"]]
            if psw not in (None, "") and psl not in (None, ""):
                ps_rows += 1
                ps_dates.append(d)

    return {
        "total_rows": total_rows,
        "completed_rows": completed_rows,
        "completed_with_pinnacle_rows": ps_rows,
        "all_dates_min": min(all_dates).isoformat() if all_dates else None,
        "all_dates_max": max(all_dates).isoformat() if all_dates else None,
        "ps_dates_min": min(ps_dates).isoformat() if ps_dates else None,
        "ps_dates_max": max(ps_dates).isoformat() if ps_dates else None,
    }


def main() -> int:
    args = parse_args()
    report_path = Path(args.report) if Path(args.report).is_absolute() else (ROOT / args.report)
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    payload, ext, source = download_season_file(args.year)
    target = DATA_DIR / f"atp-{args.year}{ext}"
    if ext != ".xlsx":
        raise RuntimeError(
            f"Downloaded tennis-data workbook is {ext}, but the current backtest/audit stack expects .xlsx. "
            f"Source={source}"
        )

    temp_path: Path | None = None
    if not args.dry_run:
        target.write_bytes(payload)
        inspect_path = target
    else:
        temp_path = _write_temp_and_return(payload, target)
        inspect_path = temp_path

    stats = inspect_xlsx(inspect_path)
    lines = [
        "Tennis-Data ATP Refresh",
        f"Generated UTC: {datetime.utcnow().isoformat()}Z",
        f"Year: {args.year}",
        f"Source: {source}",
        f"Target: {target}",
        f"Mode: {'DRY RUN' if args.dry_run else 'LIVE'}",
        "",
        f"Rows total: {stats['total_rows']}",
        f"Rows completed: {stats['completed_rows']}",
        f"Rows completed with Pinnacle PSW/PSL: {stats['completed_with_pinnacle_rows']}",
        f"All-date range: {stats['all_dates_min']} -> {stats['all_dates_max']}",
        f"PSW/PSL date range: {stats['ps_dates_min']} -> {stats['ps_dates_max']}",
    ]
    text = "\n".join(lines)
    print(text)
    if not args.dry_run:
        report_path.write_text(text + "\n", encoding="utf-8")
    if temp_path and temp_path.exists():
        temp_path.unlink(missing_ok=True)
    return 0


def _write_temp_and_return(payload: bytes, target: Path) -> Path:
    temp = target.with_name(target.stem + ".tmp.xlsx")
    temp.write_bytes(payload)
    return temp


if __name__ == "__main__":
    raise SystemExit(main())
