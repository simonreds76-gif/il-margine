#!/usr/bin/env python3
"""
Backtest fair-odds model vs tennis-data Pinnacle closing prices (ATP).

Inputs:
  - data/backtest/atp-2024.xlsx
  - data/backtest/atp-2025.xlsx
  - optional: data/backtest/atp-2026.xlsx
  - local OnCourt CSVs in data/oncourt/

Outputs:
  - data/backtest/backtest-results-YYYY.csv
  - data/backtest/backtest-log.txt
  - console summary: calibration, log loss vs Pinnacle no-vig, ROI by value threshold

Run:
  python scripts/backtest-fair-odds.py
  python scripts/backtest-fair-odds.py --include-2026
  python scripts/backtest-fair-odds.py --challenger   # Challenger from oncourt_games + data/pinnacle-odds-*.csv
  python scripts/backtest-fair-odds.py --dry-run --limit-matches 300

V2 formula (SPW-inversion fix for p_a/p_b double-counting):
  python scripts/backtest-fair-odds.py --files data/backtest/atp-2024.xlsx --v2
  # Output: data/backtest/backtest-results-2024-v2.csv

Temperature calibration (apples-to-apples ROI check):
  python scripts/backtest-fair-odds.py --files data/backtest/atp-2024.xlsx --temperature 1.18
  # Output: data/backtest/backtest-results-2024-calibrated.csv
  python scripts/edge-analysis.py --files data/backtest/backtest-results-2024-calibrated.csv
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import sys
import unicodedata
from bisect import bisect_left
from collections import Counter, defaultdict, deque
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from statistics import mean
from typing import Any, Callable

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
if str(Path(__file__).resolve().parent) not in sys.path:
    sys.path.insert(0, str(Path(__file__).resolve().parent))

from src.lib.tennis_prob import prob_match_best_of_3, prob_match_best_of_5
import sackmann_tml_id_map as idmap

try:
    from hybrid_v2 import compute_point_probs_bc
    HAS_HYBRID_V2 = True
except ImportError:
    HAS_HYBRID_V2 = False

try:
    from matchup_model import MatchupStats, compute_match_point_probs
    HAS_MATCHUP_MODEL = True
except ImportError:
    HAS_MATCHUP_MODEL = False

try:
    from fatigue_model import build_recent_matches, compute_fatigue_rust_delta
    HAS_FATIGUE_MODEL = True
except ImportError:
    HAS_FATIGUE_MODEL = False

# Core fair-odds constants (mirrors oncourt-compute-fair-odds.py where applicable)
DEFAULT_ELO = 1500.0
HYBRID_ELO_WEIGHT_DEFAULT = 0.58
HYBRID_ELO_WEIGHT_MIN_MATCHES = 28
POINT_CLAMP = (0.42, 0.80)
RETURN_CLAMP = (0.20, 0.55)
SHRINKAGE_N = 40

SURFACE_LEAGUE_AVG = {"Hard": 0.64, "Clay": 0.62, "Grass": 0.67, "I.hard": 0.64, "N/A": 0.64}
SURFACE_AVG_HOLD = {"Hard": 0.4342, "Clay": 0.4189, "Grass": 0.5053, "I.hard": 0.5099, "N/A": 0.44}
SURFACE_AVG_RETURN = {"Hard": 0.3467, "Clay": 0.3643, "Grass": 0.3166, "I.hard": 0.3413, "N/A": 0.35}

RANK_LOGIT_SCALE = 0.95
RANK_CLASS_GAP_RATIO_START = 1.40
RANK_CLASS_GAP_CAP = 0.07
RANK_POINTS_BLEND = 0.35
POINTS_LOGIT_SCALE = 1.25
POINTS_MIN_BASE = 5.0

FORM_WEIGHT = 0.07
FORM_CAP = 0.03
FORM_MIN_MATCHES = 3
FATIGUE_PLAYED_YESTERDAY = 0.008
FATIGUE_DENSE_5D = 0.007
FATIGUE_MODERATE_COMPOUND = 0.004
FATIGUE_CAP = 0.015
RUST_THRESHOLD_MODERATE = 28
RUST_THRESHOLD_SEVERE = 42
RUST_PENALTY_MODERATE = 0.008
RUST_PENALTY_SEVERE = 0.015
RUST_CAP = 0.015
FORM_TOTAL_CAP = 0.05

H2H_MIN_MATCHES = 5
H2H_WEIGHT = 0.03
H2H_CAP = 0.02

TOURNAMENT_HISTORY_MIN_MATCHES = 2
TOURNAMENT_HISTORY_WEIGHT = 0.08
TOURNAMENT_HISTORY_CAP = 0.03
TOURNAMENT_HISTORY_PRIOR_K = 3.0
TOURNAMENT_HISTORY_PRIOR_P = 0.5

YTD_FORM_MIN_MATCHES = 5
YTD_FORM_WEIGHT = 0.07
YTD_FORM_CAP = 0.03
LAST5_FORM_MIN_MATCHES = 3
LAST5_FORM_WEIGHT = 0.04
LAST5_FORM_CAP = 0.015
STREAK_SOFT_START = 3
STREAK_HARD_START = 6
STREAK_UNIT = 0.004
STREAK_CAP = 0.03
RESULTS_FORM_TOTAL_CAP = 0.055

FORM_CRISIS_LOSS_STREAK = 5
FORM_CRISIS_YTD_MIN_MATCHES = 6
FORM_CRISIS_YTD_WINRATE_MAX = 0.25
RANK_WEIGHT_CRISIS_MULT = 0.55

DELTA_CAP_HIGH = 0.07
DELTA_CAP_MEDIUM = 0.055
DELTA_CAP_LOW = 0.045

VS_LEFTIE_WEIGHT = 0.03
VS_LEFTIE_CAP = 0.015

AGE_CAP = 0.01
MISSING_ACTIVITY_BASE_PENALTY = 0.010
MISSING_ACTIVITY_AGE_START = 31
MISSING_ACTIVITY_AGE_WEIGHT = 0.0020
MISSING_ACTIVITY_PENALTY_CAP = 0.035
ELO_RELIABILITY_MIN = 0.50

# Optional CPI (surface-speed) overlay for backtest A/B versus base model.
DEFAULT_CPI_LOOKBACK_YEARS = 3
DEFAULT_CPI_MATCH_WEIGHT = 0.025
DEFAULT_CPI_MATCH_CAP = 0.012
DEFAULT_CPI_Z_CAP = 2.0
MIN_TOURNAMENT_SPEED_MATCHES = 50
TOURNAMENT_SPEED_LOOKBACK_DAYS = 365 * 3
TOURNAMENT_SPEED_VENUE_FULL_MATCHES = 150
DEFAULT_TOURNAMENT_SPEED_POINT_WEIGHT = 0.18
SPEED_RATIO_CLAMP = (0.92, 1.08)

POINT_WINDOW_12M_DAYS = 365
POINT_WINDOW_36M_DAYS = 1095
RESULTS_LOOKBACK_DAYS = 365 * 6
ELO_K_SURFACE = 24.0
ELO_K_OVERALL = 22.0

UNKNOWN_PLAYER_IDS = {3699, 3700}
DEFAULT_THRESHOLDS = (0.02, 0.05, 0.10)

# Post-hoc probability calibration (fit on 2024, validated on 2025).
# Favorite-space Platt transform:
#   q_cal = sigmoid(A + B * logit(q_raw))
# then blended by series to avoid overcorrecting strong major-event priors.
PROB_CAL_A = 0.173
PROB_CAL_B = 0.512
PROB_CAL_SERIES_BLEND = {
    "ATP250": 0.35,
    "Masters 1000": 1.00,
    "ATP500": 0.35,
    "Grand Slam": 0.00,
    "Masters Cup": 0.00,
}
SERIES_FAVORITE_PROB_CAP = {
    "ATP250": {"high": 0.89, "medium": 0.85, "low": 0.82},
}


@dataclass
class BacktestMatch:
    year: int
    date_ord: int
    date_iso: str
    tournament: str
    location: str
    tournament_key: str
    tournament_speed_key: str
    surface: str
    round_name: str
    series: str
    winner_name_short: str
    loser_name_short: str
    winner_rank: int | None
    loser_rank: int | None
    winner_pts: float | None
    loser_pts: float | None
    psw: float
    psl: float
    winner_id: int | None = None
    loser_id: int | None = None
    score: str | None = None  # e.g. "6-3 7-5" for handicap settlement


@dataclass
class HistoryEvent:
    date_ord: int
    winner_id: int
    loser_id: int
    surface: str
    tour_key: str
    tournament_speed_key: str
    winner_stats: tuple[int, ...] | None  # (svc_w, svc_t, ret_w, ret_t) or +5 decomposed
    loser_stats: tuple[int, ...] | None


def _float(v: Any, default: float | None = None) -> float | None:
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _int(v: Any, default: int | None = None) -> int | None:
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def _clamp(x: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, x))


def _sample_scale(match_count: int | None, full_matches: int) -> float:
    if match_count is None:
        return 0.0
    if full_matches <= 0:
        return 1.0
    return _clamp(float(match_count) / float(full_matches), 0.0, 1.0)


def _compute_tournament_speed_signal(
    *,
    league_avg: float,
    venue_entry: dict[str, Any] | None,
) -> tuple[float, dict[str, Any]]:
    if league_avg <= 0.0:
        return 0.0, {}

    debug = {
        "venue_spw": None,
        "venue_spw_matches": 0,
        "venue_ratio": None,
        "venue_signal": 0.0,
    }
    if not venue_entry:
        return 0.0, debug

    venue_spw = _float(venue_entry.get("venue_avg_spw"))
    venue_n = int(venue_entry.get("match_count") or 0)
    if venue_spw is None or venue_spw <= 0.0:
        return 0.0, debug

    speed_ratio_span = max(SPEED_RATIO_CLAMP[1] - 1.0, 1.0 - SPEED_RATIO_CLAMP[0], 1e-6)
    ratio = _clamp(venue_spw / league_avg, SPEED_RATIO_CLAMP[0], SPEED_RATIO_CLAMP[1])
    signal = _clamp((ratio - 1.0) / speed_ratio_span, -1.0, 1.0)
    signal *= _sample_scale(venue_n, TOURNAMENT_SPEED_VENUE_FULL_MATCHES)
    debug.update(
        {
            "venue_spw": venue_spw,
            "venue_spw_matches": venue_n,
            "venue_ratio": ratio,
            "venue_signal": signal,
        }
    )
    return signal, debug


def _logit(p: float) -> float:
    p = _clamp(p, 1e-6, 1.0 - 1e-6)
    return math.log(p / (1.0 - p))


def _sigmoid(z: float) -> float:
    if z >= 0:
        ez = math.exp(-z)
        return 1.0 / (1.0 + ez)
    ez = math.exp(z)
    return ez / (1.0 + ez)


def _safe_prob(p: float) -> float:
    return _clamp(p, 1e-6, 1.0 - 1e-6)


def _apply_temperature(p: float, T: float = 1.18) -> float:
    """Post-hoc temperature scaling: p_cal = sigmoid(logit(p) / T). T>1 reduces overconfidence."""
    p = _safe_prob(p)
    return 1.0 / (1.0 + math.exp(-_logit(p) / T))


def _mean_std(values: list[float]) -> tuple[float, float]:
    vals = [float(x) for x in values if x is not None]
    n = len(vals)
    if n == 0:
        return (0.0, 1.0)
    mu = sum(vals) / float(n)
    if n == 1:
        return (mu, 1.0)
    var = sum((x - mu) ** 2 for x in vals) / float(n)
    std = math.sqrt(var)
    if std < 1e-9:
        std = 1.0
    return (mu, std)


def _surface_candidates(surface: str) -> tuple[str, ...]:
    if surface == "I.hard":
        return ("I.hard", "Hard", "N/A")
    if surface == "Hard":
        return ("Hard", "I.hard", "N/A")
    return (surface, "N/A")


def _ascii_letters(text: str) -> str:
    s = unicodedata.normalize("NFKD", str(text or ""))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"[^a-z]", "", s.lower())


def _tour_key(name: str) -> str:
    raw = (name or "").strip().lower()
    if not raw:
        return ""
    parts = [p.strip() for p in re.split(r"\s*-\s*", raw) if p.strip()]
    core = parts[-1] if len(parts) >= 2 and len(parts[-1]) >= 4 else raw
    core = re.sub(r"\b\d{4}\b", " ", core)
    core = re.sub(r"\b(challenger|qualifiers?|qualifying|qualification|atp|wta)\b", " ", core)
    core = re.sub(r"[^a-z0-9]+", " ", core)
    return " ".join(core.split())


def _normalize_speed_fragment(text: str | None) -> str:
    raw = (text or "").strip().lower()
    if not raw:
        return ""
    raw = raw.replace("&", " and ")
    raw = raw.replace("’", "'")
    raw = re.sub(r"\bu[\.\s]*s[\.\s]*\b", "us ", raw)
    raw = re.sub(r"\bmens[' ]s\b", "mens", raw)
    raw = re.sub(r"\bwomens[' ]s\b", "womens", raw)
    raw = re.sub(r"\bmen[' ]s\b", "mens", raw)
    raw = re.sub(r"\bwomen[' ]s\b", "womens", raw)
    raw = re.sub(r"\bchampionships\b", "championship", raw)
    raw = re.sub(r"\bqualifiers?\b|\bqualifying\b|\bqualification\b", " ", raw)
    raw = re.sub(r"\b(challenger|atp|wta)\b", " ", raw)
    raw = re.sub(r"\b\d{4}\b", " ", raw)
    raw = re.sub(r"[^a-z0-9]+", " ", raw)
    return " ".join(raw.split())


def _tournament_speed_key(name: str | None, location: str | None = None) -> str:
    raw = (name or "").strip()
    loc = (location or "").strip()
    if not raw and not loc:
        return ""
    parts = [p.strip() for p in re.split(r"\s*-\s*", raw) if p.strip()]
    title = parts[0] if parts else raw
    inferred_location = parts[-1] if len(parts) >= 2 else ""
    title_key = _normalize_speed_fragment(title)
    location_key = _normalize_speed_fragment(loc or inferred_location)
    if title_key and location_key:
        return f"{title_key} {location_key}"
    return title_key or location_key


def _tour_key_candidates(*names: str) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for name in names:
        raw = (name or "").strip()
        if not raw:
            continue
        lower = raw.lower()
        parts = [p.strip() for p in re.split(r"\s*-\s*", lower) if p.strip()]
        cands = [_tour_key(lower)]
        if parts:
            cands.append(_tour_key(parts[0]))
            cands.append(_tour_key(parts[-1]))
        if len(parts) >= 2:
            cands.append(_tour_key(parts[0] + " " + parts[-1]))
        for c in cands:
            if not c or c in seen:
                continue
            seen.add(c)
            out.append(c)
    return out


def _to_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _court_to_surface(court_name: str, tennis_surface: str | None = None) -> str:
    c = (court_name or "").strip().upper()
    t = (tennis_surface or "").strip().upper()
    if "CLAY" in c or "TERRE" in c or "CLAY" in t:
        return "Clay"
    if "GRASS" in c or "GRASS" in t:
        return "Grass"
    if "INDOOR" in c and "HARD" in c:
        return "I.hard"
    if "I.HARD" in c:
        return "I.hard"
    if "HARD" in c or "DECOTURF" in c or "ACRYLIC" in c or "HARD" in t:
        return "Hard"
    if "CARPET" in c or "CARPET" in t:
        return "I.hard"
    return "N/A"


def _is_supported_tour(name: str, rank: int | None) -> bool:
    raw = (name or "")
    u = raw.upper()
    if re.search(r"\b[MW]\d{1,2}\b", raw, flags=re.IGNORECASE):
        return False
    if "ITF" in u or "FUTURES" in u:
        return False
    if rank is not None and rank <= 3:
        return True
    if "CHALLENGER" in u:
        return True
    if "ATP" in u or "MASTERS" in u or "GRAND SLAM" in u:
        return True
    return False


def _parse_tennis_data_short_name(name: str) -> dict[str, Any]:
    raw = (name or "").strip()
    if not raw:
        return {"raw": raw, "surname_norm": "", "surname_tokens": [], "abbr": "", "abbr_is_initials": False}
    parts = raw.split()
    if len(parts) == 1:
        norm = idmap.normalize_name(raw)
        toks = norm.split()
        return {
            "raw": raw,
            "surname_norm": norm,
            "surname_tokens": toks,
            "abbr": "",
            "abbr_is_initials": False,
        }

    def _is_abbrev_token(tok: str) -> bool:
        letters = _ascii_letters(tok)
        if not letters:
            return False
        if "." in tok:
            return len(letters) <= 3
        return len(letters) <= 2

    tail = []
    i = len(parts) - 1
    while i >= 1 and _is_abbrev_token(parts[i]):
        tail.append(parts[i])
        i -= 1
    if not tail:
        tail = [parts[-1]]
        i = len(parts) - 2

    surname_raw = " ".join(parts[: i + 1]) if i >= 0 else parts[0]
    given_raw = " ".join(reversed(tail))
    surname_norm = idmap.normalize_name(surname_raw)
    surname_tokens = surname_norm.split()
    abbr = _ascii_letters(given_raw)
    abbr_is_initials = (len(tail) >= 2 and len(abbr) >= 2) or (given_raw.count(".") >= 2 and len(abbr) >= 2)
    return {
        "raw": raw,
        "surname_norm": surname_norm,
        "surname_tokens": surname_tokens,
        "abbr": abbr,
        "abbr_is_initials": abbr_is_initials,
    }


def _player_name_meta(name: str) -> dict[str, str]:
    norm = idmap.normalize_name(name)
    toks = norm.split()
    given = toks[:-1] if len(toks) >= 2 else toks
    initials = "".join(t[:1] for t in given if t)
    return {
        "norm": norm,
        "first_token": given[0] if given else "",
        "initials": initials,
    }


def _filter_candidates_with_abbrev(
    candidates: set[int],
    parsed: dict[str, Any],
    player_meta: dict[int, dict[str, str]],
) -> set[int]:
    if len(candidates) <= 1:
        return candidates

    abbr = parsed.get("abbr") or ""
    if not abbr:
        return candidates

    filtered = set()
    if parsed.get("abbr_is_initials"):
        for pid in candidates:
            ini = (player_meta.get(pid, {}).get("initials") or "")
            if ini.startswith(abbr):
                filtered.add(pid)
    else:
        for pid in candidates:
            mt = player_meta.get(pid, {})
            first_token = mt.get("first_token") or ""
            ini = mt.get("initials") or ""
            if first_token.startswith(abbr) or ini.startswith(abbr):
                filtered.add(pid)
    return filtered if filtered else candidates


def _resolve_short_name_to_oncourt_id(
    short_name: str,
    indexes: tuple[dict[str, dict[str, set[int]]], dict[int, set[str]], dict[int, dict[str, Any]]],
    player_meta: dict[int, dict[str, str]],
    popularity_by_pid: dict[int, int] | None = None,
) -> tuple[int | None, str | None]:
    parsed = _parse_tennis_data_short_name(short_name)
    surname_norm = parsed.get("surname_norm") or ""
    surname_tokens = parsed.get("surname_tokens") or []
    abbr = parsed.get("abbr") or ""

    candidates_to_try: list[str] = []
    if surname_norm and abbr:
        candidates_to_try.append(f"{abbr[:1]} {surname_norm}")
        if len(abbr) > 1:
            candidates_to_try.append(f"{abbr} {surname_norm}")
            if parsed.get("abbr_is_initials"):
                candidates_to_try.append(f"{' '.join(abbr)} {surname_norm}")
    candidates_to_try.append(short_name)

    tried = set()
    safe_methods = {
        "full",
        "first_last",
        "initial_last",
        "first_last_compact",
        "last_only",
        "last_first",
        "last_first_cross",
        "first_hyphen_space",
        "first_hyphen_compact",
        "hyphen_space",
        "hyphen_compact",
        "fuzzy_singleton",
    }
    for cand in candidates_to_try:
        key = idmap.normalize_name(cand)
        if not key or key in tried:
            continue
        tried.add(key)
        pid, method = idmap._resolve_oncourt_id(cand, indexes, source_meta=None)
        base_method = (method or "").split("_meta_")[0]
        if pid is not None and base_method in safe_methods:
            return pid, f"{method}_short"

    indexes_map = indexes[0]
    weak = set()

    if abbr and surname_tokens:
        first = abbr[:1]
        last = surname_tokens[-1]
        weak |= set(indexes_map["initial_last"].get(f"{first} {last}", set()))
        weak |= set(indexes_map["last_only"].get(last, set()))
        weak |= set(indexes_map["last_first"].get(f"{last} {first}", set()))
        if len(surname_tokens) >= 2:
            last_two = " ".join(surname_tokens[-2:])
            weak |= set(indexes_map["last_two"].get(last_two, set()))
            weak |= set(indexes_map["first_last_two"].get(f"{first} {last_two}", set()))

    weak = _filter_candidates_with_abbrev(weak, parsed, player_meta)
    if len(weak) == 1:
        return next(iter(weak)), "short_custom"

    if len(weak) > 1:
        norm_short = idmap.normalize_name(short_name)
        fuzzy = []
        pid_to_norm = indexes[1]
        for pid in weak:
            for target in pid_to_norm.get(pid, set()):
                if idmap._is_close_name(norm_short, target, max_dist=2):
                    fuzzy.append(pid)
                    break
        fuzzy = sorted(set(fuzzy))
        if len(fuzzy) == 1:
            return fuzzy[0], "short_fuzzy"

    if len(weak) > 1 and popularity_by_pid:
        ranked = sorted(
            ((popularity_by_pid.get(pid, 0), pid) for pid in weak),
            reverse=True,
        )
        top_count, top_pid = ranked[0]
        second_count = ranked[1][0] if len(ranked) > 1 else -1
        if top_count >= 25 and (second_count <= 0 or top_count >= int(second_count * 1.5)):
            return top_pid, "short_popularity"

    return None, None


def _build_tennis_data_name_map(
    short_names: set[str],
    oncourt_players: list[dict[str, Any]],
    popularity_by_pid: dict[int, int] | None = None,
) -> tuple[dict[str, int], dict[str, Any]]:
    singles_players = [r for r in oncourt_players if "/" not in (r.get("name") or "")]
    indexes = idmap._build_oncourt_indexes(singles_players)
    player_meta = {int(row["id"]): _player_name_meta(row.get("name") or "") for row in singles_players if row.get("id") is not None}

    resolved: dict[str, int] = {}
    methods = Counter()
    unresolved = []

    for short_name in sorted(short_names):
        pid, method = _resolve_short_name_to_oncourt_id(
            short_name,
            indexes,
            player_meta,
            popularity_by_pid=popularity_by_pid,
        )
        if pid is None:
            unresolved.append(short_name)
            continue
        resolved[short_name] = int(pid)
        methods[method or "unknown"] += 1

    report = {
        "source_names": len(short_names),
        "mapped_names": len(resolved),
        "unmapped_names": len(unresolved),
        "oncourt_index_players": len(singles_players),
        "method_counts": dict(methods),
        "unmapped_examples": unresolved[:200],
    }
    return resolved, report


def _empty_hist() -> dict[str, list[int]]:
    return {
        "d": [],
        "svc_w": [0], "svc_t": [0], "ret_w": [0], "ret_t": [0],
        "first_in": [0], "first_attempts": [0], "first_won": [0],
        "second_attempts": [0], "second_won": [0],
        "ret_vs_first_won": [0], "ret_vs_first_faced": [0],
        "ret_vs_second_won": [0], "ret_vs_second_faced": [0],
    }


def _hist_add(
    hist: dict[str, list[int]],
    d: int,
    svc_w: int,
    svc_t: int,
    ret_w: int,
    ret_t: int,
    first_in: int = 0,
    first_attempts: int = 0,
    first_won: int = 0,
    second_attempts: int = 0,
    second_won: int = 0,
    ret_vs_first_won: int = 0,
    ret_vs_first_faced: int = 0,
    ret_vs_second_won: int = 0,
    ret_vs_second_faced: int = 0,
) -> None:
    hist["d"].append(d)
    hist["svc_w"].append(hist["svc_w"][-1] + svc_w)
    hist["svc_t"].append(hist["svc_t"][-1] + svc_t)
    hist["ret_w"].append(hist["ret_w"][-1] + ret_w)
    hist["ret_t"].append(hist["ret_t"][-1] + ret_t)
    hist["first_in"].append(hist["first_in"][-1] + first_in)
    hist["first_attempts"].append(hist["first_attempts"][-1] + first_attempts)
    hist["first_won"].append(hist["first_won"][-1] + first_won)
    hist["second_attempts"].append(hist["second_attempts"][-1] + second_attempts)
    hist["second_won"].append(hist["second_won"][-1] + second_won)
    hist["ret_vs_first_won"].append(hist["ret_vs_first_won"][-1] + ret_vs_first_won)
    hist["ret_vs_first_faced"].append(hist["ret_vs_first_faced"][-1] + ret_vs_first_faced)
    hist["ret_vs_second_won"].append(hist["ret_vs_second_won"][-1] + ret_vs_second_won)
    hist["ret_vs_second_faced"].append(hist["ret_vs_second_faced"][-1] + ret_vs_second_faced)


def _hist_query(hist: dict[str, list[int]] | None, date_ord: int, window_days: int) -> tuple[int, int, int, int, int, int, int, int, int, int, int, int, int, int]:
    """Returns (n, svc_w, svc_t, ret_w, ret_t, first_in, first_attempts, first_won, second_attempts, second_won,
    ret_vs_first_won, ret_vs_first_faced, ret_vs_second_won, ret_vs_second_faced)."""
    if not hist:
        return (0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
    dates = hist["d"]
    if not dates:
        return (0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
    hi = bisect_left(dates, date_ord)
    lo = bisect_left(dates, date_ord - window_days, 0, hi)
    if hi <= lo:
        return (0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
    n = hi - lo
    svc_w = hist["svc_w"][hi] - hist["svc_w"][lo]
    svc_t = hist["svc_t"][hi] - hist["svc_t"][lo]
    ret_w = hist["ret_w"][hi] - hist["ret_w"][lo]
    ret_t = hist["ret_t"][hi] - hist["ret_t"][lo]
    first_in = hist["first_in"][hi] - hist["first_in"][lo]
    first_attempts = hist["first_attempts"][hi] - hist["first_attempts"][lo]
    first_won = hist["first_won"][hi] - hist["first_won"][lo]
    second_attempts = hist["second_attempts"][hi] - hist["second_attempts"][lo]
    second_won = hist["second_won"][hi] - hist["second_won"][lo]
    rvf_w = hist["ret_vs_first_won"][hi] - hist["ret_vs_first_won"][lo]
    rvf_f = hist["ret_vs_first_faced"][hi] - hist["ret_vs_first_faced"][lo]
    rvs_w = hist["ret_vs_second_won"][hi] - hist["ret_vs_second_won"][lo]
    rvs_f = hist["ret_vs_second_faced"][hi] - hist["ret_vs_second_faced"][lo]
    return (n, svc_w, svc_t, ret_w, ret_t, first_in, first_attempts, first_won, second_attempts, second_won,
            rvf_w, rvf_f, rvs_w, rvs_f)


class ModelState:
    def __init__(self) -> None:
        self.elo_surface: dict[tuple[int, str], float] = {}
        self.elo_overall: dict[int, float] = {}
        self.elo_surface_matches: Counter[tuple[int, str]] = Counter()
        self.elo_overall_matches: Counter[int] = Counter()

        self.point_hist: dict[tuple[int, str], dict[str, list[int]]] = defaultdict(_empty_hist)

        self.result_dates: dict[int, list[int]] = defaultdict(list)
        self.result_wins_prefix: dict[int, list[int]] = defaultdict(lambda: [0])
        self.result_opp_elo_prefix: dict[int, list[float]] = defaultdict(lambda: [0.0])
        self.result_outcomes: dict[int, list[int]] = defaultdict(list)

        self.tour_history: dict[tuple[int, str], list[int]] = defaultdict(lambda: [0, 0])  # wins, matches
        self.tournament_speed_hist: dict[tuple[str, str], dict[str, list[float] | list[int]]] = defaultdict(
            lambda: {"dates": [], "serve_won_prefix": [0.0], "serve_total_prefix": [0.0], "match_prefix": [0]}
        )
        self.h2h: dict[tuple[int, int, str], list[int]] = defaultdict(lambda: [0, 0, 0])  # wins_a, wins_b, matches
        self.vs_leftie: dict[tuple[int, str], list[int]] = defaultdict(lambda: [0, 0])  # wins, matches

    def _elo_expected(self, elo_a: float, elo_b: float) -> float:
        return 1.0 / (1.0 + 10.0 ** ((elo_b - elo_a) / 400.0))

    def update(self, ev: HistoryEvent, lefties: set[int]) -> None:
        w = ev.winner_id
        l = ev.loser_id
        surf = ev.surface

        e_w_s = self.elo_surface.get((w, surf), DEFAULT_ELO)
        e_l_s = self.elo_surface.get((l, surf), DEFAULT_ELO)
        exp_w_s = self._elo_expected(e_w_s, e_l_s)
        self.elo_surface[(w, surf)] = e_w_s + ELO_K_SURFACE * (1.0 - exp_w_s)
        self.elo_surface[(l, surf)] = e_l_s + ELO_K_SURFACE * (0.0 - (1.0 - exp_w_s))
        self.elo_surface_matches[(w, surf)] += 1
        self.elo_surface_matches[(l, surf)] += 1

        e_w_o = self.elo_overall.get(w, DEFAULT_ELO)
        e_l_o = self.elo_overall.get(l, DEFAULT_ELO)
        exp_w_o = self._elo_expected(e_w_o, e_l_o)
        self.elo_overall[w] = e_w_o + ELO_K_OVERALL * (1.0 - exp_w_o)
        self.elo_overall[l] = e_l_o + ELO_K_OVERALL * (0.0 - (1.0 - exp_w_o))
        self.elo_overall_matches[w] += 1
        self.elo_overall_matches[l] += 1

        self._append_result(w, ev.date_ord, 1, e_l_s)
        self._append_result(l, ev.date_ord, 0, e_w_s)

        if ev.tour_key:
            tr_w = self.tour_history[(w, ev.tour_key)]
            tr_w[0] += 1
            tr_w[1] += 1
            tr_l = self.tour_history[(l, ev.tour_key)]
            tr_l[1] += 1
            if ev.winner_stats and ev.loser_stats:
                ws = ev.winner_stats
                ls = ev.loser_stats
                serve_won = float(ws[0] + ls[0])
                if len(ws) >= 6 and len(ls) >= 6:
                    serve_total = float(ws[5] + ls[5])
                else:
                    serve_total = float(ws[1] + ls[1])
                if serve_total > 0 and ev.tournament_speed_key:
                    hist = self.tournament_speed_hist[(ev.tournament_speed_key, surf)]
                    dates = hist["dates"]
                    won_prefix = hist["serve_won_prefix"]
                    total_prefix = hist["serve_total_prefix"]
                    match_prefix = hist["match_prefix"]
                    dates.append(ev.date_ord)
                    won_prefix.append(won_prefix[-1] + serve_won)
                    total_prefix.append(total_prefix[-1] + serve_total)
                    match_prefix.append(match_prefix[-1] + 1)

        a, b = (w, l) if w < l else (l, w)
        for s in (surf, "N/A"):
            rec = self.h2h[(a, b, s)]
            if w == a:
                rec[0] += 1
            else:
                rec[1] += 1
            rec[2] += 1

        if ev.winner_stats:
            ws = ev.winner_stats
            dec = (ws[4], ws[5], ws[6], ws[7], ws[8]) if len(ws) >= 9 else (0, 0, 0, 0, 0)
            # Winner's return vs opponent (loser) serve: loser's first_in - first_won, etc.
            ls = ev.loser_stats
            if ls and len(ls) >= 9:
                rvf_w, rvf_f = max(0, ls[4] - ls[6]), ls[4]
                rvs_w, rvs_f = max(0, ls[7] - ls[8]), ls[7]
            else:
                rvf_w, rvf_f, rvs_w, rvs_f = 0, 0, 0, 0
            _hist_add(self.point_hist[(w, surf)], ev.date_ord, ws[0], ws[1], ws[2], ws[3], *dec, rvf_w, rvf_f, rvs_w, rvs_f)
            _hist_add(self.point_hist[(w, "N/A")], ev.date_ord, ws[0], ws[1], ws[2], ws[3], *dec, rvf_w, rvf_f, rvs_w, rvs_f)
        if ev.loser_stats:
            ls = ev.loser_stats
            dec = (ls[4], ls[5], ls[6], ls[7], ls[8]) if len(ls) >= 9 else (0, 0, 0, 0, 0)
            # Loser's return vs opponent (winner) serve
            ws = ev.winner_stats
            if ws and len(ws) >= 9:
                rvf_w, rvf_f = max(0, ws[4] - ws[6]), ws[4]
                rvs_w, rvs_f = max(0, ws[7] - ws[8]), ws[7]
            else:
                rvf_w, rvf_f, rvs_w, rvs_f = 0, 0, 0, 0
            _hist_add(self.point_hist[(l, surf)], ev.date_ord, ls[0], ls[1], ls[2], ls[3], *dec, rvf_w, rvf_f, rvs_w, rvs_f)
            _hist_add(self.point_hist[(l, "N/A")], ev.date_ord, ls[0], ls[1], ls[2], ls[3], *dec, rvf_w, rvf_f, rvs_w, rvs_f)

        if l in lefties:
            vl = self.vs_leftie[(w, surf)]
            vl[0] += 1
            vl[1] += 1
            vla = self.vs_leftie[(w, "N/A")]
            vla[0] += 1
            vla[1] += 1
        if w in lefties:
            vl = self.vs_leftie[(l, surf)]
            vl[1] += 1
            vla = self.vs_leftie[(l, "N/A")]
            vla[1] += 1

    def _append_result(self, pid: int, d: int, is_win: int, opp_elo_surface: float) -> None:
        self.result_dates[pid].append(d)
        self.result_outcomes[pid].append(is_win)
        self.result_wins_prefix[pid].append(self.result_wins_prefix[pid][-1] + is_win)
        self.result_opp_elo_prefix[pid].append(self.result_opp_elo_prefix[pid][-1] + float(opp_elo_surface))

    def player_stats(self, pid: int, surface: str, date_ord: int) -> dict[str, Any]:
        for s in _surface_candidates(surface):
            hist = self.point_hist.get((pid, s))
            q12 = _hist_query(hist, date_ord, POINT_WINDOW_12M_DAYS)
            q36 = _hist_query(hist, date_ord, POINT_WINDOW_36M_DAYS)
            n12, svc_w12, svc_t12, ret_w12, ret_t12 = q12[0], q12[1], q12[2], q12[3], q12[4]
            n36, svc_w36, svc_t36, ret_w36, ret_t36 = q36[0], q36[1], q36[2], q36[3], q36[4]
            fi12, fa12, fw12, sa12, sw12 = q12[5], q12[6], q12[7], q12[8], q12[9]
            fi36, fa36, fw36, sa36, sw36 = q36[5], q36[6], q36[7], q36[8], q36[9]
            rvf_w12, rvf_f12, rvs_w12, rvs_f12 = q12[10], q12[11], q12[12], q12[13]
            rvf_w36, rvf_f36, rvs_w36, rvs_f36 = q36[10], q36[11], q36[12], q36[13]
            if n12 > 0 or n36 > 0:
                out: dict[str, Any] = {
                    "surface_used": s,
                    "hold_pct": (svc_w12 / svc_t12) if svc_t12 > 0 else None,
                    "return_pct": (ret_w12 / ret_t12) if ret_t12 > 0 else None,
                    "hold_pct_long": (svc_w36 / svc_t36) if svc_t36 > 0 else None,
                    "return_pct_long": (ret_w36 / ret_t36) if ret_t36 > 0 else None,
                    "match_count": n12,
                    "service_pts": svc_t12,
                    "return_pts": ret_t12,
                }
                if fa12 > 0:
                    out["first_serve_pct"] = fi12 / fa12
                    out["first_serve_win_pct"] = fw12 / fi12 if fi12 > 0 else None
                if sa12 > 0:
                    out["second_serve_win_pct"] = sw12 / sa12
                if fa36 > 0:
                    out["first_serve_pct_long"] = fi36 / fa36
                    out["first_serve_win_pct_long"] = fw36 / fi36 if fi36 > 0 else None
                if sa36 > 0:
                    out["second_serve_win_pct_long"] = sw36 / sa36
                if rvf_f12 > 0:
                    out["ret_vs_first_win_pct"] = rvf_w12 / rvf_f12
                if rvs_f12 > 0:
                    out["ret_vs_second_win_pct"] = rvs_w12 / rvs_f12
                if rvf_f36 > 0:
                    out["ret_vs_first_win_pct_long"] = rvf_w36 / rvf_f36
                if rvs_f36 > 0:
                    out["ret_vs_second_win_pct_long"] = rvs_w36 / rvs_f36
                return out
        return {
            "surface_used": None,
            "hold_pct": None,
            "return_pct": None,
            "hold_pct_long": None,
            "return_pct_long": None,
            "match_count": 0,
            "service_pts": 0,
            "return_pts": 0,
        }

    def recent_activity(self, pid: int, date_ord: int) -> dict[str, Any] | None:
        dates = self.result_dates.get(pid)
        if not dates:
            return None
        idx = bisect_left(dates, date_ord)
        if idx <= 0:
            return None

        i21 = bisect_left(dates, date_ord - 21, 0, idx)
        i5 = bisect_left(dates, date_ord - 5, 0, idx)
        iy0 = bisect_left(dates, date_ord - 1, 0, idx)
        iy1 = bisect_left(dates, date_ord, 0, idx)

        wins_pref = self.result_wins_prefix[pid]
        opp_pref = self.result_opp_elo_prefix[pid]

        m21 = idx - i21
        w21 = wins_pref[idx] - wins_pref[i21]
        m5 = idx - i5
        avg_opp = (opp_pref[idx] - opp_pref[i21]) / m21 if m21 > 0 else None
        return {
            "matches_last_21d": m21,
            "wins_last_21d": w21,
            "win_rate_21d": (w21 / m21) if m21 > 0 else None,
            "matches_last_5d": m5,
            "played_yesterday": iy1 > iy0,
            "last_match_ord": dates[idx - 1],
            "avg_opponent_elo": avg_opp,
        }

    def results_snapshot(self, pid: int, date_ord: int) -> dict[str, Any]:
        dates = self.result_dates.get(pid)
        if not dates:
            return {
                "matches_total": 0,
                "ytd_matches": 0,
                "ytd_wins": 0,
                "ytd_win_rate": None,
                "last5_matches": 0,
                "last5_wins": 0,
                "last5_win_rate": None,
                "win_streak": 0,
                "loss_streak": 0,
            }

        idx = bisect_left(dates, date_ord)
        if idx <= 0:
            return {
                "matches_total": 0,
                "ytd_matches": 0,
                "ytd_wins": 0,
                "ytd_win_rate": None,
                "last5_matches": 0,
                "last5_wins": 0,
                "last5_win_rate": None,
                "win_streak": 0,
                "loss_streak": 0,
            }

        today = date.fromordinal(date_ord)
        ytd_start = date(today.year, 1, 1).toordinal()
        i_ytd = bisect_left(dates, ytd_start, 0, idx)
        wins_pref = self.result_wins_prefix[pid]
        outcomes = self.result_outcomes[pid]

        ytd_n = idx - i_ytd
        ytd_w = wins_pref[idx] - wins_pref[i_ytd]

        i_last5 = max(0, idx - 5)
        l5_n = idx - i_last5
        l5_w = wins_pref[idx] - wins_pref[i_last5]

        win_streak = 0
        loss_streak = 0
        if idx > 0:
            first = outcomes[idx - 1]
            streak = 0
            j = idx - 1
            while j >= 0 and outcomes[j] == first:
                streak += 1
                j -= 1
            if first == 1:
                win_streak = streak
            else:
                loss_streak = streak

        return {
            "matches_total": idx,
            "ytd_matches": ytd_n,
            "ytd_wins": ytd_w,
            "ytd_win_rate": (ytd_w / ytd_n) if ytd_n > 0 else None,
            "last5_matches": l5_n,
            "last5_wins": l5_w,
            "last5_win_rate": (l5_w / l5_n) if l5_n > 0 else None,
            "win_streak": win_streak,
            "loss_streak": loss_streak,
        }

    def tour_record(self, pid: int, tour_key: str) -> tuple[int, int]:
        rec = self.tour_history.get((pid, tour_key))
        if not rec:
            return (0, 0)
        return (int(rec[0]), int(rec[1]))

    def tournament_speed_profile(
        self,
        tournament_speed_key: str,
        surface: str,
        date_ord: int,
        *,
        lookback_days: int = TOURNAMENT_SPEED_LOOKBACK_DAYS,
        min_matches: int = MIN_TOURNAMENT_SPEED_MATCHES,
    ) -> dict[str, Any] | None:
        if not tournament_speed_key:
            return None
        for s in _surface_candidates(surface):
            hist = self.tournament_speed_hist.get((tournament_speed_key, s))
            if not hist:
                continue
            dates = hist["dates"]
            hi = bisect_left(dates, date_ord)
            lo = bisect_left(dates, date_ord - lookback_days, 0, hi)
            matches = hist["match_prefix"][hi] - hist["match_prefix"][lo]
            serve_total = hist["serve_total_prefix"][hi] - hist["serve_total_prefix"][lo]
            if matches < min_matches or serve_total <= 0:
                continue
            serve_won = hist["serve_won_prefix"][hi] - hist["serve_won_prefix"][lo]
            return {
                "surface_used": s,
                "match_count": int(matches),
                "venue_avg_spw": serve_won / serve_total,
            }
        return None

    def h2h_record(self, p1: int, p2: int, surface: str) -> tuple[int, int, int, int, int] | None:
        a, b = (p1, p2) if p1 < p2 else (p2, p1)
        for s in _surface_candidates(surface):
            rec = self.h2h.get((a, b, s))
            if rec:
                return (a, b, int(rec[0]), int(rec[1]), int(rec[2]))
        return None

    def vs_leftie_win_pct(self, pid: int, surface: str) -> float:
        for s in _surface_candidates(surface):
            rec = self.vs_leftie.get((pid, s))
            if rec and rec[1] > 0:
                return rec[0] / rec[1]
        return 0.5


def _age_years_at(birthdate_str: str | None, on_date: date) -> float | None:
    if not birthdate_str:
        return None
    try:
        bd = date.fromisoformat(str(birthdate_str)[:10])
    except (ValueError, TypeError):
        return None
    return (on_date - bd).days / 365.25


def _age_factor_at(birthdate_str: str | None, on_date: date) -> float:
    age = _age_years_at(birthdate_str, on_date)
    if age is None:
        return 0.0
    if age >= 30:
        return -0.004 * (age - 30)
    if age < 22:
        return 0.002 * (22 - age)
    return 0.0


def _results_form_component(snapshot: dict[str, Any]) -> float:
    d = 0.0
    ytd_n = int(snapshot.get("ytd_matches") or 0)
    ytd_wr = snapshot.get("ytd_win_rate")
    if ytd_n >= YTD_FORM_MIN_MATCHES and ytd_wr is not None:
        ytd_scale = min(1.0, ytd_n / 14.0)
        d += _clamp((float(ytd_wr) - 0.5) * YTD_FORM_WEIGHT * ytd_scale, -YTD_FORM_CAP, YTD_FORM_CAP)

    l5_n = int(snapshot.get("last5_matches") or 0)
    l5_wr = snapshot.get("last5_win_rate")
    if l5_n >= LAST5_FORM_MIN_MATCHES and l5_wr is not None:
        l5_scale = min(1.0, l5_n / 5.0)
        d += _clamp((float(l5_wr) - 0.5) * LAST5_FORM_WEIGHT * l5_scale, -LAST5_FORM_CAP, LAST5_FORM_CAP)

    loss_streak = int(snapshot.get("loss_streak") or 0)
    win_streak = int(snapshot.get("win_streak") or 0)
    if loss_streak >= STREAK_SOFT_START:
        streak_pen = (loss_streak - STREAK_SOFT_START + 1) * STREAK_UNIT
        if loss_streak >= STREAK_HARD_START:
            streak_pen += 0.006
        d -= min(STREAK_CAP, streak_pen)
    elif win_streak >= STREAK_SOFT_START:
        streak_boost = (win_streak - STREAK_SOFT_START + 1) * STREAK_UNIT
        if win_streak >= STREAK_HARD_START:
            streak_boost += 0.004
        d += min(STREAK_CAP, streak_boost)
    return _clamp(d, -RESULTS_FORM_TOTAL_CAP, RESULTS_FORM_TOTAL_CAP)


def _is_form_crisis(snapshot: dict[str, Any]) -> bool:
    ytd_n = int(snapshot.get("ytd_matches") or 0)
    ytd_wr = snapshot.get("ytd_win_rate")
    loss_streak = int(snapshot.get("loss_streak") or 0)
    if loss_streak >= FORM_CRISIS_LOSS_STREAK:
        return True
    if ytd_n >= FORM_CRISIS_YTD_MIN_MATCHES and ytd_wr is not None and float(ytd_wr) <= FORM_CRISIS_YTD_WINRATE_MAX:
        return True
    return False


def _recent_reliability(activity: dict[str, Any] | None, birthdate: str | None, match_date: date) -> float:
    if not activity:
        age = _age_years_at(birthdate, match_date)
        rel = 0.72
        if age is not None and age >= 31:
            rel -= min(0.24, (age - 31) * 0.025)
        return max(ELO_RELIABILITY_MIN, rel)

    m21 = int(activity.get("matches_last_21d") or 0)
    if m21 >= 8:
        return 1.00
    if m21 >= 5:
        return 0.95
    if m21 >= 3:
        return 0.90
    if m21 >= 1:
        return 0.84

    last_ord = activity.get("last_match_ord")
    if last_ord:
        days = match_date.toordinal() - int(last_ord)
        if days >= 90:
            return 0.55
        if days >= 60:
            return 0.65
        if days >= 35:
            return 0.75
    return 0.80


def _form_component_only(
    activity: dict[str, Any] | None,
    player_elo_surface: float,
    birthdate: str | None,
    match_date: date,
) -> float:
    """Form only: outperformance vs expected given opponent quality. No fatigue/rust."""
    if not activity:
        age = _age_years_at(birthdate, match_date)
        penalty = MISSING_ACTIVITY_BASE_PENALTY
        if age is not None and age >= MISSING_ACTIVITY_AGE_START:
            penalty += (age - MISSING_ACTIVITY_AGE_START) * MISSING_ACTIVITY_AGE_WEIGHT
        if age is not None and age >= 34:
            penalty += min(0.010, (age - 34) * 0.002)
        return -min(MISSING_ACTIVITY_PENALTY_CAP, penalty)

    m21 = int(activity.get("matches_last_21d") or 0)
    wr = activity.get("win_rate_21d")
    avg_opp = _float(activity.get("avg_opponent_elo"), 1500.0)

    delta_form = 0.0
    if m21 >= FORM_MIN_MATCHES and wr is not None and avg_opp is not None:
        expected_wr = 1.0 / (1.0 + 10.0 ** ((avg_opp - player_elo_surface) / 400.0))
        delta_form = (float(wr) - expected_wr) * FORM_WEIGHT
        delta_form = _clamp(delta_form, -FORM_CAP, FORM_CAP)
    return delta_form


def _form_fatigue_rust(
    activity: dict[str, Any] | None,
    player_elo_surface: float,
    birthdate: str | None,
    match_date: date,
) -> float:
    """Legacy: form + V4 fatigue + rust. Used when fatigue model not available."""
    delta_form = _form_component_only(activity, player_elo_surface, birthdate, match_date)
    if not activity:
        return delta_form

    delta_fatigue = 0.0
    if activity.get("played_yesterday"):
        delta_fatigue -= FATIGUE_PLAYED_YESTERDAY
    m5 = int(activity.get("matches_last_5d") or 0)
    if m5 >= 3:
        delta_fatigue -= FATIGUE_DENSE_5D
    elif m5 >= 2 and activity.get("played_yesterday"):
        delta_fatigue -= FATIGUE_MODERATE_COMPOUND
    delta_fatigue = max(-FATIGUE_CAP, delta_fatigue)

    delta_rust = 0.0
    last_ord = activity.get("last_match_ord")
    if last_ord:
        days = match_date.toordinal() - int(last_ord)
        if days >= RUST_THRESHOLD_SEVERE:
            delta_rust = -RUST_PENALTY_SEVERE
        elif days >= RUST_THRESHOLD_MODERATE:
            delta_rust = -RUST_PENALTY_MODERATE
        delta_rust = max(-RUST_CAP, delta_rust)

    return delta_form + delta_fatigue + delta_rust


def _compute_tour_history_component(state: ModelState, pid: int, tour_key: str) -> float:
    if not tour_key:
        return 0.0
    wins, matches = state.tour_record(pid, tour_key)
    if matches < TOURNAMENT_HISTORY_MIN_MATCHES:
        return 0.0
    wr = (wins + TOURNAMENT_HISTORY_PRIOR_P * TOURNAMENT_HISTORY_PRIOR_K) / (matches + TOURNAMENT_HISTORY_PRIOR_K)
    scale = min(1.0, matches / 8.0)
    return _clamp((wr - 0.5) * TOURNAMENT_HISTORY_WEIGHT * scale, -TOURNAMENT_HISTORY_CAP, TOURNAMENT_HISTORY_CAP)


def _series_calibration_blend(series: str, surface: str, confidence: str) -> float:
    blend = PROB_CAL_SERIES_BLEND.get(series, 1.0)
    if series in ("ATP250", "ATP500"):
        if confidence == "medium":
            blend *= 0.55
        elif confidence == "low":
            blend *= 0.35
        if series == "ATP500" and surface == "Hard":
            blend *= 0.00
        elif surface in ("Clay", "Grass"):
            blend *= 0.75
    return _clamp(blend, 0.0, 1.0)


def _series_rank_weight_multiplier(series: str, confidence: str) -> float:
    if series == "ATP250":
        return {"high": 0.90, "medium": 0.72, "low": 0.60}.get(confidence, 1.0)
    if series == "ATP500":
        return {"high": 0.95, "medium": 0.78, "low": 0.65}.get(confidence, 1.0)
    return 1.0


def _series_delta_multipliers(series: str, surface: str, confidence: str) -> tuple[float, float, float, float]:
    if series == "ATP250":
        if confidence == "high":
            return (1.12, 1.15, 0.88, 0.90)
        if confidence == "medium":
            return (1.18, 1.22, 0.76, 0.82)
        if confidence == "low":
            return (1.24, 1.28, 0.66, 0.72)
    if series == "ATP500":
        if surface == "Hard":
            if confidence == "high":
                return (0.98, 1.02, 0.68, 0.80)
            if confidence == "medium":
                return (1.02, 1.06, 0.60, 0.72)
            if confidence == "low":
                return (1.06, 1.10, 0.52, 0.66)
        if confidence == "high":
            return (1.02, 1.06, 0.78, 0.88)
        if confidence == "medium":
            return (1.08, 1.12, 0.70, 0.82)
        if confidence == "low":
            return (1.14, 1.18, 0.62, 0.74)
    return (1.0, 1.0, 1.0, 1.0)


def _series_crisis_rank_multiplier(series: str) -> float:
    if series == "ATP250":
        return 0.42
    if series == "ATP500":
        return 0.48
    return RANK_WEIGHT_CRISIS_MULT


def _apply_series_probability_guard(p1_win: float, series: str, surface: str, confidence: str) -> float:
    caps = SERIES_FAVORITE_PROB_CAP.get(series)
    if not caps:
        return _safe_prob(p1_win)
    cap = float(caps.get(confidence, 0.90))
    if series == "ATP500" and surface == "Hard":
        cap -= 0.02
    cap = _clamp(cap, 0.78, 0.93)
    p = _safe_prob(p1_win)
    fav_is_p1 = p >= 0.5
    q = p if fav_is_p1 else (1.0 - p)
    q = min(q, cap)
    return q if fav_is_p1 else (1.0 - q)


def _calibrate_match_prob(p1_win: float, series: str, surface: str, confidence: str) -> float:
    p = _safe_prob(p1_win)
    fav_is_p1 = p >= 0.5
    q_raw = p if fav_is_p1 else (1.0 - p)
    q_raw = _safe_prob(q_raw)
    z = PROB_CAL_A + PROB_CAL_B * _logit(q_raw)
    q_cal = _sigmoid(z)
    blend = _series_calibration_blend(series, surface, confidence)
    q = (1.0 - blend) * q_raw + blend * q_cal
    q = _safe_prob(q)
    return q if fav_is_p1 else (1.0 - q)


def _compute_match_probability(
    match: BacktestMatch,
    state: ModelState,
    birthdate_by_player: dict[int, str | None],
    lefties: set[int],
    *,
    use_v2_formula: bool = False,
    use_matchup_model: bool = True,
    use_decomposed_return: bool = True,
    cpi_enabled: bool = False,
    cpi_lookup: dict[tuple[int, str, str], float] | None = None,
    cpi_year_surface_stats: dict[tuple[int, str], tuple[float, float]] | None = None,
    cpi_surface_stats: dict[str, tuple[float, float]] | None = None,
    cpi_fallback_years: int = DEFAULT_CPI_LOOKBACK_YEARS,
    cpi_match_weight: float = DEFAULT_CPI_MATCH_WEIGHT,
    cpi_match_cap: float = DEFAULT_CPI_MATCH_CAP,
    cpi_z_cap: float = DEFAULT_CPI_Z_CAP,
    tournament_speed_enabled: bool = True,
    tournament_speed_point_weight: float = DEFAULT_TOURNAMENT_SPEED_POINT_WEIGHT,
    games_by_player: dict[int, list[dict[str, Any]]] | None = None,
    tour_surface_map: dict[int, str] | None = None,
) -> tuple[float, str, dict[str, Any]]:
    p1 = int(match.winner_id)  # player1 is row winner (actual outcome)
    p2 = int(match.loser_id)
    surface = match.surface
    match_dt = date.fromordinal(match.date_ord)

    s1 = state.player_stats(p1, surface, match.date_ord)
    s2 = state.player_stats(p2, surface, match.date_ord)

    has_s1 = int(s1.get("match_count") or 0) > 0
    has_s2 = int(s2.get("match_count") or 0) > 0

    surf_hold = SURFACE_AVG_HOLD.get(surface, SURFACE_AVG_HOLD["N/A"])
    surf_ret = SURFACE_AVG_RETURN.get(surface, SURFACE_AVG_RETURN["N/A"])
    hold_shift = SURFACE_LEAGUE_AVG.get(surface, SURFACE_LEAGUE_AVG["N/A"]) - surf_hold
    ret_shift = (1.0 - SURFACE_LEAGUE_AVG.get(surface, SURFACE_LEAGUE_AVG["N/A"])) - surf_ret

    h1_12 = _float(s1.get("hold_pct"), surf_hold) or surf_hold
    r1_12 = _float(s1.get("return_pct"), surf_ret) or surf_ret
    h2_12 = _float(s2.get("hold_pct"), surf_hold) or surf_hold
    r2_12 = _float(s2.get("return_pct"), surf_ret) or surf_ret
    h1_long = _float(s1.get("hold_pct_long"), h1_12) or h1_12
    r1_long = _float(s1.get("return_pct_long"), r1_12) or r1_12
    h2_long = _float(s2.get("hold_pct_long"), h2_12) or h2_12
    r2_long = _float(s2.get("return_pct_long"), r2_12) or r2_12

    mc1_12 = int(s1.get("match_count") or 0)
    mc2_12 = int(s2.get("match_count") or 0)
    min_matches_12 = min(mc1_12, mc2_12)

    if min_matches_12 >= 25:
        recent_weight = 0.75
    elif min_matches_12 >= 15:
        recent_weight = 0.6
    elif min_matches_12 >= 8:
        recent_weight = 0.5
    else:
        recent_weight = 0.3
    if surface == "Clay":
        recent_weight = min(recent_weight, 0.6)
    elif surface == "Grass":
        recent_weight = max(recent_weight, 0.55)

    hold1_raw = recent_weight * h1_12 + (1.0 - recent_weight) * h1_long
    ret1_raw = recent_weight * r1_12 + (1.0 - recent_weight) * r1_long
    hold2_raw = recent_weight * h2_12 + (1.0 - recent_weight) * h2_long
    ret2_raw = recent_weight * r2_12 + (1.0 - recent_weight) * r2_long

    alpha1 = mc1_12 / (mc1_12 + SHRINKAGE_N) if mc1_12 else 0.0
    alpha2 = mc2_12 / (mc2_12 + SHRINKAGE_N) if mc2_12 else 0.0
    hold1 = alpha1 * hold1_raw + (1.0 - alpha1) * surf_hold if alpha1 > 0 else surf_hold
    ret1 = alpha1 * ret1_raw + (1.0 - alpha1) * surf_ret if alpha1 > 0 else surf_ret
    hold2 = alpha2 * hold2_raw + (1.0 - alpha2) * surf_hold if alpha2 > 0 else surf_hold
    ret2 = alpha2 * ret2_raw + (1.0 - alpha2) * surf_ret if alpha2 > 0 else surf_ret

    hold1_eff = _clamp(hold1 + hold_shift, POINT_CLAMP[0], POINT_CLAMP[1])
    hold2_eff = _clamp(hold2 + hold_shift, POINT_CLAMP[0], POINT_CLAMP[1])
    ret1_eff = _clamp(ret1 + ret_shift, RETURN_CLAMP[0], RETURN_CLAMP[1])
    ret2_eff = _clamp(ret2 + ret_shift, RETURN_CLAMP[0], RETURN_CLAMP[1])

    league_avg = SURFACE_LEAGUE_AVG.get(surface, SURFACE_LEAGUE_AVG["N/A"])
    if league_avg <= 0.0 or league_avg >= 1.0:
        league_avg = 0.64
    tour_speed_signal = 0.0
    tour_speed_multiplier = 0.0
    tour_speed_debug: dict[str, Any] = {}
    hold1_model = hold1_eff
    hold2_model = hold2_eff
    ret1_model = ret1_eff
    ret2_model = ret2_eff
    if tournament_speed_enabled and tournament_speed_point_weight > 0.0:
        venue_entry = state.tournament_speed_profile(match.tournament_speed_key, surface, match.date_ord)
        tour_speed_signal, tour_speed_debug = _compute_tournament_speed_signal(
            league_avg=league_avg,
            venue_entry=venue_entry,
        )
        if abs(tour_speed_signal) > 1e-9:
            tour_speed_multiplier = _clamp(tour_speed_signal * tournament_speed_point_weight, -0.35, 0.35)
            return_baseline = 1.0 - league_avg
            hold1_model = _clamp(league_avg + (hold1_eff - league_avg) * (1.0 + tour_speed_multiplier), POINT_CLAMP[0], POINT_CLAMP[1])
            hold2_model = _clamp(league_avg + (hold2_eff - league_avg) * (1.0 + tour_speed_multiplier), POINT_CLAMP[0], POINT_CLAMP[1])
            ret1_model = _clamp(return_baseline + (ret1_eff - return_baseline) * (1.0 - tour_speed_multiplier), RETURN_CLAMP[0], RETURN_CLAMP[1])
            ret2_model = _clamp(return_baseline + (ret2_eff - return_baseline) * (1.0 - tour_speed_multiplier), RETURN_CLAMP[0], RETURN_CLAMP[1])

    matchup_method = "legacy"
    if use_matchup_model and HAS_MATCHUP_MODEL and (s1 or s2):
        def _blend(s: dict[str, Any], k12: str, klong: str) -> float | None:
            v12 = _float(s.get(k12))
            vlong = _float(s.get(klong), v12)
            if v12 is None:
                return vlong
            if vlong is None:
                return v12
            return recent_weight * v12 + (1.0 - recent_weight) * vlong

        def _build_matchup_row(s: dict[str, Any], hold_val: float, ret_val: float) -> dict[str, Any]:
            row: dict[str, Any] = {
                "hold_pct": hold_val,
                "return_pct": ret_val,
                "match_count": int(s.get("match_count") or 0),
                "service_pts": int(s.get("service_pts") or 0),
            }
            if s.get("first_serve_pct") is not None:
                row["first_serve_pct"] = _blend(s, "first_serve_pct", "first_serve_pct_long")
                row["first_serve_win_pct"] = _blend(s, "first_serve_win_pct", "first_serve_win_pct_long")
                row["second_serve_win_pct"] = _blend(s, "second_serve_win_pct", "second_serve_win_pct_long")
            if use_decomposed_return and s.get("ret_vs_first_win_pct") is not None:
                row["ret_vs_first_win_pct"] = _blend(s, "ret_vs_first_win_pct", "ret_vs_first_win_pct_long")
            if use_decomposed_return and s.get("ret_vs_second_win_pct") is not None:
                row["ret_vs_second_win_pct"] = _blend(s, "ret_vs_second_win_pct", "ret_vs_second_win_pct_long")
            return row

        # Use venue-speed adjusted SPW/RPW values so matchup branch is consistent with
        # the fallback branch and live fair-odds pipeline.
        row1 = _build_matchup_row(s1 or {}, hold1_model, ret1_model)
        row2 = _build_matchup_row(s2 or {}, hold2_model, ret2_model)
        stats_a = MatchupStats.from_db_row(row1)
        stats_b = MatchupStats.from_db_row(row2)
        p_a, p_b, _ = compute_match_point_probs(stats_a, stats_b, surface)
        # Classify for decomposed-vs-fallback diagnostic
        if stats_a.has_decomposed and stats_b.has_decomposed:
            matchup_method = "both_decomposed"
        elif stats_a.has_decomposed or stats_b.has_decomposed:
            matchup_method = "one_decomposed"
        else:
            matchup_method = "neither_decomposed"
    elif use_v2_formula and HAS_HYBRID_V2:
        p_a, p_b = compute_point_probs_bc(
            hold1_model, ret1_model, hold2_model, ret2_model,
            surface, league_avg=SURFACE_LEAGUE_AVG
        )
    else:
        p_a = _clamp((hold1_model * (1.0 - ret2_model)) / league_avg, POINT_CLAMP[0], POINT_CLAMP[1])
        p_b = _clamp((hold2_model * (1.0 - ret1_model)) / league_avg, POINT_CLAMP[0], POINT_CLAMP[1])
    is_best_of_5 = (match.series or "").strip() == "Grand Slam"
    p_serve_return = prob_match_best_of_5(p_a, p_b) if is_best_of_5 else prob_match_best_of_3(p_a, p_b)

    e1_s = state.elo_surface.get((p1, surface), DEFAULT_ELO)
    e2_s = state.elo_surface.get((p2, surface), DEFAULT_ELO)
    e1_o = state.elo_overall.get(p1)
    e2_o = state.elo_overall.get(p2)
    e1_is_real = ((p1, surface) in state.elo_surface) or (p1 in state.elo_overall)
    e2_is_real = ((p2, surface) in state.elo_surface) or (p2 in state.elo_overall)

    act1 = state.recent_activity(p1, match.date_ord)
    act2 = state.recent_activity(p2, match.date_ord)
    rel1 = _recent_reliability(act1, birthdate_by_player.get(p1), match_dt)
    rel2 = _recent_reliability(act2, birthdate_by_player.get(p2), match_dt)
    elo_reliability = 0.5 * (rel1 + rel2)

    e1_s_eff = DEFAULT_ELO + rel1 * (e1_s - DEFAULT_ELO)
    e2_s_eff = DEFAULT_ELO + rel2 * (e2_s - DEFAULT_ELO)
    p_elo_surface = 1.0 / (1.0 + 10.0 ** ((e2_s_eff - e1_s_eff) / 400.0))
    if e1_o is not None and e2_o is not None:
        e1_o_eff = DEFAULT_ELO + rel1 * (float(e1_o) - DEFAULT_ELO)
        e2_o_eff = DEFAULT_ELO + rel2 * (float(e2_o) - DEFAULT_ELO)
        p_elo_overall = 1.0 / (1.0 + 10.0 ** ((e2_o_eff - e1_o_eff) / 400.0))
        p_elo = 0.5 * p_elo_surface + 0.5 * p_elo_overall
    else:
        p_elo = p_elo_surface

    r1 = match.winner_rank
    r2 = match.loser_rank
    pts1 = match.winner_pts
    pts2 = match.loser_pts

    p_rank = None
    p_rank_atp = None
    p_rank_points = None

    if r1 is not None and r2 is not None and r1 > 0 and r2 > 0:
        log_ratio = math.log(max(1.0, float(r2)) / max(1.0, float(r1)))
        p_rank_atp = _sigmoid(log_ratio / RANK_LOGIT_SCALE)
        rr = max(r1, r2) / float(min(r1, r2))
        if rr >= RANK_CLASS_GAP_RATIO_START:
            extra = min(RANK_CLASS_GAP_CAP, math.log(rr / RANK_CLASS_GAP_RATIO_START) * 0.08)
            p_rank_atp = _clamp(p_rank_atp + (extra if r1 < r2 else -extra), 0.02, 0.98)

    if pts1 is not None and pts2 is not None and (pts1 > 0 or pts2 > 0):
        log_pts_ratio = math.log((max(0.0, pts1) + POINTS_MIN_BASE) / (max(0.0, pts2) + POINTS_MIN_BASE))
        p_rank_points = _sigmoid(log_pts_ratio / POINTS_LOGIT_SCALE)

    if p_rank_atp is not None and p_rank_points is not None:
        p_rank = (1.0 - RANK_POINTS_BLEND) * p_rank_atp + RANK_POINTS_BLEND * p_rank_points
    elif p_rank_atp is not None:
        p_rank = p_rank_atp
    elif p_rank_points is not None:
        p_rank = p_rank_points

    both_elo_default = not e1_is_real and not e2_is_real
    if both_elo_default and p_rank is None and not has_s1 and not has_s2:
        confidence = "none"
    elif not has_s1 and not has_s2:
        confidence = "low"
    elif not has_s1 or not has_s2:
        confidence = "low" if min_matches_12 < 5 else "medium"
    elif min_matches_12 < 5:
        confidence = "medium"
    elif min_matches_12 >= 10:
        confidence = "high"
    else:
        confidence = "medium"
    if confidence in ("high", "medium") and (act1 is None or act2 is None):
        confidence = "medium" if confidence == "high" else "low"
    form_mult, results_mult, structural_mult, cap_mult = _series_delta_multipliers(match.series, surface, confidence)

    cpi_value = None
    cpi_year = None
    cpi_z = 0.0
    if cpi_enabled and cpi_lookup:
        season_year = match_dt.year
        surf = "Hard" if surface == "I.hard" else surface
        key_candidates = _tour_key_candidates(
            match.tournament_key,
            match.tournament,
            f"{match.tournament} - {match.location}".strip(" -"),
        )
        if surf in ("Hard", "Clay", "Grass") and key_candidates:
            for yr_back in range(0, max(0, int(cpi_fallback_years)) + 1):
                y = season_year - yr_back
                found = None
                for tk in key_candidates:
                    found = cpi_lookup.get((y, surf, tk))
                    if found is not None:
                        cpi_value = float(found)
                        cpi_year = y
                        break
                if cpi_value is not None:
                    break
        if cpi_value is not None and cpi_year is not None:
            ys_stats = cpi_year_surface_stats or {}
            s_stats = cpi_surface_stats or {}
            mean_std = ys_stats.get((cpi_year, surf), s_stats.get(surf, (0.0, 1.0)))
            mu, sd = float(mean_std[0]), float(mean_std[1])
            if sd <= 1e-9:
                sd = 1.0
            cpi_z = _clamp((cpi_value - mu) / sd, -abs(cpi_z_cap), abs(cpi_z_cap))

    p1_results = state.results_snapshot(p1, match.date_ord)
    p2_results = state.results_snapshot(p2, match.date_ord)
    p1_crisis = _is_form_crisis(p1_results)
    p2_crisis = _is_form_crisis(p2_results)

    serve_rel = min(1.0, min_matches_12 / float(HYBRID_ELO_WEIGHT_MIN_MATCHES)) if (has_s1 and has_s2) else 0.0
    w_sr = (1.0 - HYBRID_ELO_WEIGHT_DEFAULT) * (0.35 + 0.65 * serve_rel) if (has_s1 and has_s2) else 0.0
    w_elo = HYBRID_ELO_WEIGHT_DEFAULT * (0.65 + 0.35 * elo_reliability)
    w_rank = 0.0
    rank_gap_strength = 0.0
    if p_rank is not None:
        if r1 is not None and r2 is not None and r1 > 0 and r2 > 0:
            rr = max(r1, r2) / float(min(r1, r2))
            rank_strength = _clamp((rr - 1.0) / 3.0, 0.25, 1.0)
        else:
            rank_strength = 0.45
        rank_gap_strength = _clamp(abs(p_rank - 0.5) * 2.0, 0.0, 1.0)
        w_rank = 0.18 + 0.34 * rank_strength + 0.10 * rank_gap_strength
        if (p_rank - 0.5) * (p_elo - 0.5) < 0:
            w_elo *= max(0.45, 1.0 - 0.75 * rank_gap_strength)
            w_sr *= max(0.70, 1.0 - 0.35 * rank_gap_strength)
            w_rank += 0.18 * rank_gap_strength
        w_rank *= _series_rank_weight_multiplier(match.series, confidence)

    if not has_s1 or not has_s2:
        w_sr *= 0.15
        w_elo += 0.08
        if p_rank is not None:
            w_rank += 0.12
    if both_elo_default:
        w_elo *= 0.30
        if p_rank is not None:
            w_rank += 0.35
    if p_rank is not None and w_rank > 0:
        crisis_rank_mult = _series_crisis_rank_multiplier(match.series)
        if p_rank > 0.5 and p1_crisis and not p2_crisis:
            w_rank *= crisis_rank_mult
        elif p_rank < 0.5 and p2_crisis and not p1_crisis:
            w_rank *= crisis_rank_mult

    total_w = w_sr + w_elo + w_rank
    if total_w <= 0:
        p1_win = 0.5
    else:
        sr_weight = w_sr / total_w
        elo_weight = w_elo / total_w
        rank_weight = w_rank / total_w
        z = sr_weight * _logit(p_serve_return) + elo_weight * _logit(p_elo)
        if p_rank is not None and rank_weight > 0:
            z += rank_weight * _logit(p_rank)
        p1_win = _sigmoid(z)

    delta_p1 = 0.0
    h2h_row = state.h2h_record(p1, p2, surface)
    if h2h_row:
        a, _b, wins_a, wins_b, h2h_n = h2h_row
        if h2h_n >= H2H_MIN_MATCHES:
            wins_p1 = wins_a if p1 == a else wins_b
            win_rate_p1 = wins_p1 / h2h_n
            delta_p1 += structural_mult * _clamp((win_rate_p1 - 0.5) * H2H_WEIGHT, -H2H_CAP, H2H_CAP)

    if p2 in lefties:
        delta_p1 += structural_mult * (state.vs_leftie_win_pct(p1, surface) - 0.5) * VS_LEFTIE_WEIGHT
    if p1 in lefties:
        delta_p1 -= structural_mult * (state.vs_leftie_win_pct(p2, surface) - 0.5) * VS_LEFTIE_WEIGHT

    cpi_delta = 0.0
    if cpi_enabled and cpi_value is not None and cpi_match_weight > 0.0:
        # Same conservative style-speed interaction as live fair-odds.
        style_edge = (hold1 - hold2) - (ret1 - ret2)
        cpi_delta = _clamp(cpi_z * style_edge * cpi_match_weight, -cpi_match_cap, cpi_match_cap)
        delta_p1 += structural_mult * cpi_delta

    if HAS_FATIGUE_MODEL and games_by_player is not None and tour_surface_map is not None:
        p1_form = _form_component_only(act1, e1_s, birthdate_by_player.get(p1), match_dt)
        p2_form = _form_component_only(act2, e2_s, birthdate_by_player.get(p2), match_dt)
        delta_form_part = form_mult * 0.5 * (p1_form - p2_form)
        p1_recent = build_recent_matches(
            p1, games_by_player.get(p1, []), tour_surface_map, match_dt, lookback_days=21
        )
        p2_recent = build_recent_matches(
            p2, games_by_player.get(p2, []), tour_surface_map, match_dt, lookback_days=21
        )
        delta_fatigue_rust, _ = compute_fatigue_rust_delta(
            p1_recent, p2_recent, match_dt, surface, tour_id=None, debug=False
        )
        delta_form = _clamp(
            delta_form_part + structural_mult * delta_fatigue_rust,
            -FORM_TOTAL_CAP,
            FORM_TOTAL_CAP,
        )
    else:
        p1_form = _form_fatigue_rust(act1, e1_s, birthdate_by_player.get(p1), match_dt)
        p2_form = _form_fatigue_rust(act2, e2_s, birthdate_by_player.get(p2), match_dt)
        delta_form = _clamp(form_mult * 0.5 * (p1_form - p2_form), -FORM_TOTAL_CAP, FORM_TOTAL_CAP)
    delta_p1 += delta_form

    p1_rf = _results_form_component(p1_results)
    p2_rf = _results_form_component(p2_results)
    delta_results = _clamp(results_mult * (p1_rf - p2_rf), -RESULTS_FORM_TOTAL_CAP, RESULTS_FORM_TOTAL_CAP)
    delta_p1 += delta_results

    t1 = _compute_tour_history_component(state, p1, match.tournament_key)
    t2 = _compute_tour_history_component(state, p2, match.tournament_key)
    delta_tour = structural_mult * _clamp(t1 - t2, -TOURNAMENT_HISTORY_CAP, TOURNAMENT_HISTORY_CAP)
    delta_p1 += delta_tour

    delta_p1 += 0.5 * (_age_factor_at(birthdate_by_player.get(p1), match_dt) - _age_factor_at(birthdate_by_player.get(p2), match_dt))

    cap_components = [
        VS_LEFTIE_CAP,
        cpi_match_cap if cpi_enabled else 0.0,
        AGE_CAP,
        FORM_TOTAL_CAP,
        RESULTS_FORM_TOTAL_CAP,
        TOURNAMENT_HISTORY_CAP,
        H2H_CAP,
    ]
    delta_p1 = _clamp(delta_p1, -max(cap_components), max(cap_components))
    overall_cap = DELTA_CAP_HIGH if confidence == "high" else (DELTA_CAP_MEDIUM if confidence == "medium" else DELTA_CAP_LOW)
    overall_cap *= cap_mult
    delta_p1 = _clamp(delta_p1, -overall_cap, overall_cap)

    p1_raw = _safe_prob(p1_win + delta_p1)
    p1_win = _calibrate_match_prob(p1_raw, match.series, surface, confidence)
    p1_win = _apply_series_probability_guard(p1_win, match.series, surface, confidence)
    debug = {
        "confidence": confidence,
        "p_serve_return": p_serve_return,
        "p_elo": p_elo,
        "p_rank": p_rank,
        "p_raw": p1_raw,
        "p_a": p_a,
        "p_b": p_b,
        "cpi_value": cpi_value,
        "cpi_year": cpi_year,
        "cpi_z": cpi_z,
        "cpi_delta": cpi_delta,
        "tournament_speed_signal": tour_speed_signal,
        "tournament_speed_multiplier": tour_speed_multiplier,
        "tournament_speed_venue_spw": tour_speed_debug.get("venue_spw"),
        "tournament_speed_match_count": tour_speed_debug.get("venue_spw_matches"),
        "matchup_method": matchup_method,
    }
    return p1_win, confidence, debug


def _normalize_name_for_match(name: str) -> str:
    """Normalize for Challenger Pinnacle vs OnCourt name matching."""
    s = (name or "").strip().lower()
    s = " ".join(s.split())
    return s


def _load_challenger_matches(
    games_path: Path,
    tours_path: Path,
    players_path: Path,
    courts_path: Path,
    pinnacle_data_dir: Path,
) -> tuple[list[BacktestMatch], Counter, set[str]]:
    """Load Challenger matches from oncourt_games + local pinnacle-odds CSV files.
    Only includes matches where we have both a result and Pinnacle odds for that date."""
    tours = _load_tour_info(tours_path, courts_path)
    challenger_tour_ids = {tid for tid, t in tours.items() if "CHALLENGER" in (t.get("name") or "").upper()}
    if not challenger_tour_ids:
        return [], Counter({"no_challenger_tours": 1}), set()

    players_by_id: dict[int, str] = {}
    with open(players_path, "r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            pid = _int(row.get("id"))
            if pid is not None:
                players_by_id[pid] = (row.get("name") or "").strip()

    # Build pinnacle lookup: (capture_date, norm_p1, norm_p2) -> (odds1, odds2)
    pinnacle_lookup: dict[tuple[str, str, str], tuple[float, float]] = {}
    pinnacle_dates: set[str] = set()
    pinnacle_files = sorted(pinnacle_data_dir.glob("pinnacle-odds-*.csv"))
    for p in pinnacle_files:
        m = re.search(r"(\d{4}-\d{2}-\d{2})", p.name)
        capture_date = m.group(1) if m else ""
        if not capture_date:
            continue
        with open(p, "r", encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                if str(row.get("league") or "").strip() != "Challenger":
                    continue
                p1 = _normalize_name_for_match(row.get("player1_name"))
                p2 = _normalize_name_for_match(row.get("player2_name"))
                if not p1 or not p2:
                    continue
                o1 = _float(row.get("odds1"))
                o2 = _float(row.get("odds2"))
                if o1 is None or o2 is None or o1 <= 1.0 or o2 <= 1.0:
                    continue
                key = (capture_date, p1, p2)
                pinnacle_lookup[key] = (float(o1), float(o2))
                pinnacle_dates.add(capture_date)

    if not pinnacle_dates:
        return [], Counter({"no_pinnacle_challenger_odds": 1}), set()

    matches: list[BacktestMatch] = []
    short_names = set()
    skip = Counter()

    with open(games_path, "r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            tid = _int(row.get("tour_id"))
            if tid not in challenger_tour_ids:
                continue
            date_str = (row.get("date") or "").strip()[:10]
            if not date_str or date_str not in pinnacle_dates:
                skip["date_not_in_pinnacle"] += 1
                continue
            wid = _int(row.get("winner_id"))
            lid = _int(row.get("loser_id"))
            if wid is None or lid is None:
                skip["missing_ids"] += 1
                continue
            winner_name = players_by_id.get(wid, "")
            loser_name = players_by_id.get(lid, "")
            if not winner_name or not loser_name:
                skip["unknown_player"] += 1
                continue
            if "/" in winner_name or "/" in loser_name:
                skip["doubles"] += 1
                continue

            dt = _to_date(date_str)
            if dt is None:
                skip["bad_date"] += 1
                continue

            wn = _normalize_name_for_match(winner_name)
            ln = _normalize_name_for_match(loser_name)
            key1 = (date_str, wn, ln)
            key2 = (date_str, ln, wn)
            if key1 in pinnacle_lookup:
                psw, psl = pinnacle_lookup[key1]
            elif key2 in pinnacle_lookup:
                psl, psw = pinnacle_lookup[key2]
            else:
                skip["no_pinnacle_odds"] += 1
                continue

            tour_info = tours.get(tid, {})
            surface = tour_info.get("surface", "Hard")
            tournament = tour_info.get("name", "Challenger")

            bt = BacktestMatch(
                year=dt.year,
                date_ord=dt.toordinal(),
                date_iso=date_str,
                tournament=tournament,
                location="",
                tournament_key=_tour_key(tournament),
                tournament_speed_key=_tournament_speed_key(tournament),
                surface=surface,
                round_name="",
                series="Challenger",
                winner_name_short=winner_name,
                loser_name_short=loser_name,
                winner_rank=None,
                loser_rank=None,
                winner_pts=None,
                loser_pts=None,
                psw=psw,
                psl=psl,
                winner_id=wid,
                loser_id=lid,
                score=(row.get("result") or "").strip() or None,
            )
            matches.append(bt)
            short_names.add(winner_name)
            short_names.add(loser_name)

    matches.sort(key=lambda x: (x.date_ord, x.tournament, x.winner_name_short, x.loser_name_short))
    return matches, skip, short_names


def _load_backtest_matches(paths: list[Path]) -> tuple[list[BacktestMatch], Counter, set[str]]:
    matches: list[BacktestMatch] = []
    short_names = set()
    skip = Counter()

    for path in paths:
        if not path.exists():
            raise FileNotFoundError(f"Missing file: {path}")
        m = re.search(r"(\d{4})", path.name)
        file_year = int(m.group(1)) if m else 0

        wb = load_workbook(path, read_only=True, data_only=True)
        required = ["Date", "Tournament", "Surface", "Court", "Round", "Series", "Winner", "Loser", "PSW", "PSL", "Comment"]
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            try:
                first_row = next(rows)
            except StopIteration:
                continue
            header = [str(h) if h is not None else "" for h in first_row]
            idx = {k: i for i, k in enumerate(header)}
            if any(key not in idx for key in required):
                continue  # skip sheets without required columns (e.g. notes)
            for row in rows:
                comment = str(row[idx["Comment"]] or "").strip()
                if comment != "Completed":
                    skip["non_completed"] += 1
                    continue

                dt = _to_date(row[idx["Date"]])
                if dt is None:
                    skip["missing_date"] += 1
                    continue
                winner = str(row[idx["Winner"]] or "").strip()
                loser = str(row[idx["Loser"]] or "").strip()
                if not winner or not loser:
                    skip["missing_name"] += 1
                    continue

                psw = _float(row[idx["PSW"]])
                psl = _float(row[idx["PSL"]])
                if psw is None or psl is None:
                    skip["missing_pinnacle_odds"] += 1
                    continue
                if psw <= 1.0 or psl <= 1.0:
                    skip["bad_pinnacle_odds"] += 1
                    continue

                surface = _court_to_surface(str(row[idx["Court"]] or ""), str(row[idx["Surface"]] or ""))
                tournament = str(row[idx["Tournament"]] or "").strip()
                location = str(row[idx.get("Location", -1)] or "").strip() if "Location" in idx else ""
                tournament_key = _tour_key(f"{tournament} - {location}".strip(" -"))
                if not tournament_key:
                    tournament_key = _tour_key(tournament)
                tournament_speed_key = _tournament_speed_key(tournament, location)

                winner_rank = _int(row[idx.get("WRank")]) if "WRank" in idx else None
                loser_rank = _int(row[idx.get("LRank")]) if "LRank" in idx else None
                winner_pts = _float(row[idx.get("WPts")]) if "WPts" in idx else None
                loser_pts = _float(row[idx.get("LPts")]) if "LPts" in idx else None

                score_str = None
                if all(k in idx for k in ["W1", "L1", "W2", "L2"]):
                    w1 = _int(row[idx["W1"]])
                    l1 = _int(row[idx["L1"]])
                    w2 = _int(row[idx["W2"]])
                    l2 = _int(row[idx["L2"]])
                    if w1 is not None and l1 is not None and w2 is not None and l2 is not None:
                        parts = [f"{w1}-{l1}", f"{w2}-{l2}"]
                        if "W3" in idx and "L3" in idx:
                            w3 = _int(row[idx["W3"]])
                            l3 = _int(row[idx["L3"]])
                            if w3 is not None and l3 is not None:
                                parts.append(f"{w3}-{l3}")
                        score_str = " ".join(parts)

                bt = BacktestMatch(
                    year=file_year,
                    date_ord=dt.toordinal(),
                    date_iso=dt.isoformat(),
                    tournament=tournament,
                    location=location,
                    tournament_key=tournament_key,
                    tournament_speed_key=tournament_speed_key,
                    surface=surface,
                    round_name=str(row[idx["Round"]] or "").strip(),
                    series=str(row[idx["Series"]] or "").strip(),
                    winner_name_short=winner,
                    loser_name_short=loser,
                    winner_rank=winner_rank,
                    loser_rank=loser_rank,
                    winner_pts=winner_pts,
                    loser_pts=loser_pts,
                    psw=float(psw),
                    psl=float(psl),
                    score=score_str,
                )
                matches.append(bt)
                short_names.add(winner)
                short_names.add(loser)

    matches.sort(key=lambda x: (x.date_ord, x.year, x.tournament, x.winner_name_short, x.loser_name_short))
    return matches, skip, short_names


def _load_oncourt_players(path: Path) -> list[dict[str, Any]]:
    rows = []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            pid = _int(row.get("id"))
            if pid is None:
                continue
            rows.append(
                {
                    "id": pid,
                    "name": (row.get("name") or "").strip(),
                    "birthdate": (row.get("birthdate") or "").strip()[:10] or None,
                    "country": (row.get("country") or "").strip().upper() or None,
                }
            )
    return rows


def _load_lefties(path: Path, valid_player_ids: set[int] | None = None) -> set[int]:
    out = set()
    if not path.exists():
        return out
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            pid = _int(row.get("player_id"))
            if pid is None:
                continue
            if valid_player_ids is not None and pid not in valid_player_ids:
                continue
            if str(row.get("cat1") or "").strip().lower() == "true":
                out.add(pid)
    return out


def _load_cpi_lookup(
    path: Path,
) -> tuple[dict[tuple[int, str, str], float], dict[tuple[int, str], tuple[float, float]], dict[str, tuple[float, float]]]:
    lookup: dict[tuple[int, str, str], float] = {}
    by_year_surface: dict[tuple[int, str], list[float]] = defaultdict(list)
    by_surface: dict[str, list[float]] = defaultdict(list)
    if not path.exists():
        return lookup, {}, {}
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            y = _int(row.get("season_year"))
            surf_raw = str(row.get("surface") or "").strip()
            key = str(row.get("tournament_key") or "").strip()
            cpi = _float(row.get("cpi"))
            if y is None or cpi is None or not key or not surf_raw:
                continue
            surf = "Hard" if surf_raw == "I.hard" else surf_raw
            if surf not in ("Hard", "Clay", "Grass"):
                continue
            idx = (int(y), surf, key)
            # Keep first if duplicate key appears.
            if idx not in lookup:
                lookup[idx] = float(cpi)
                by_year_surface[(int(y), surf)].append(float(cpi))
                by_surface[surf].append(float(cpi))
    stats_year = {k: _mean_std(v) for k, v in by_year_surface.items()}
    stats_surface = {k: _mean_std(v) for k, v in by_surface.items()}
    return lookup, stats_year, stats_surface


def _load_tour_info(tours_path: Path, courts_path: Path) -> dict[int, dict[str, Any]]:
    courts = {}
    with open(courts_path, "r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            cid = _int(row.get("id"))
            if cid is not None:
                courts[cid] = (row.get("name") or "").strip()

    out = {}
    with open(tours_path, "r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            tid = _int(row.get("id"))
            if tid is None:
                continue
            name = (row.get("name") or "").strip()
            rank = _int(row.get("rank"))
            d = _to_date(row.get("date"))
            cid = _int(row.get("court_id"))
            surface = _court_to_surface(courts.get(cid, "N/A"), None)
            out[tid] = {
                "name": name,
                "rank": rank,
                "date_ord": d.toordinal() if d else None,
                "surface": surface,
                "tour_key": _tour_key(name),
                "tournament_speed_key": _tournament_speed_key(name),
                "is_supported": _is_supported_tour(name, rank),
            }
    return out


def _load_games_for_fatigue(
    games_path: Path,
    tours: dict[int, dict[str, Any]],
    max_date_ord: int,
    player_ids: set[int] | None = None,
) -> tuple[dict[int, list[dict[str, Any]]], dict[int, str]]:
    """Load games from games_atp.csv, index by player, for fatigue model.
    If player_ids is set, only load games for those players (saves memory)."""
    tour_surface_map = {tid: str(t.get("surface") or "Hard") for tid, t in tours.items()}
    games_by_player: dict[int, list[dict[str, Any]]] = defaultdict(list)

    with open(games_path, "r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            w = _int(row.get("winner_id"))
            l = _int(row.get("loser_id"))
            t = _int(row.get("tour_id"))
            if w is None or l is None:
                continue
            if player_ids is not None and w not in player_ids and l not in player_ids:
                continue
            gd = _to_date(row.get("date"))
            if not gd:
                continue
            if gd.toordinal() >= max_date_ord:
                continue
            rec = {
                "winner_id": w,
                "loser_id": l,
                "tour_id": t,
                "round_id": _int(row.get("round_id")),
                "result": (row.get("result") or row.get("score") or "").strip(),
                "date": gd,
            }
            games_by_player[w].append(rec)
            games_by_player[l].append(rec)

    for pid in games_by_player:
        games_by_player[pid].sort(key=lambda g: g["date"], reverse=True)
    return dict(games_by_player), tour_surface_map


def _load_player_popularity(
    games_path: Path,
    players_by_id: dict[int, dict[str, Any]],
    tours: dict[int, dict[str, Any]],
) -> dict[int, int]:
    pop = Counter()
    with open(games_path, "r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            w = _int(row.get("winner_id"))
            l = _int(row.get("loser_id"))
            t = _int(row.get("tour_id"))
            if w is None or l is None or t is None:
                continue
            tour = tours.get(t)
            if not tour or not tour.get("is_supported"):
                continue
            pw = players_by_id.get(w)
            pl = players_by_id.get(l)
            if not pw or not pl:
                continue
            if "/" in (pw.get("name") or "") or "/" in (pl.get("name") or ""):
                continue
            pop[w] += 1
            pop[l] += 1
    return dict(pop)


def _load_stat_map(path: Path) -> dict[tuple[int, int, int, int], deque[tuple[int, ...]]]:
    stat_map: dict[tuple[int, int, int, int], deque[tuple[int, ...]]] = defaultdict(deque)
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            w = _int(row.get("winner_id"))
            l = _int(row.get("loser_id"))
            t = _int(row.get("tour_id"))
            rd = _int(row.get("round_id"), 0)
            if w is None or l is None or t is None:
                continue
            key = (w, l, t, int(rd or 0))
            w_fs = int(_int(row.get("w_fs"), 0) or 0)  # first serves in (OnCourt)
            l_fs = int(_int(row.get("l_fs"), 0) or 0)
            stat_map[key].append(
                (
                    int(_int(row.get("w_w1s"), 0) or 0),
                    int(_int(row.get("w_w2s"), 0) or 0),
                    int(_int(row.get("w_fsof"), 0) or 0),
                    int(_int(row.get("w_w2sof"), 0) or 0),
                    int(_int(row.get("w_rpw"), 0) or 0),
                    int(_int(row.get("w_rpwof"), 0) or 0),
                    int(_int(row.get("l_w1s"), 0) or 0),
                    int(_int(row.get("l_w2s"), 0) or 0),
                    int(_int(row.get("l_fsof"), 0) or 0),
                    int(_int(row.get("l_w2sof"), 0) or 0),
                    int(_int(row.get("l_rpw"), 0) or 0),
                    int(_int(row.get("l_rpwof"), 0) or 0),
                    w_fs,
                    l_fs,
                )
            )
    return stat_map


def _history_from_oncourt(
    games_path: Path,
    players_by_id: dict[int, dict[str, Any]],
    tours: dict[int, dict[str, Any]],
    stat_map: dict[tuple[int, int, int, int], deque[tuple[int, ...]]],
    max_date_ord: int,
) -> tuple[list[HistoryEvent], Counter]:
    events: list[HistoryEvent] = []
    skip = Counter()

    with open(games_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            w = _int(row.get("winner_id"))
            l = _int(row.get("loser_id"))
            t = _int(row.get("tour_id"))
            rd = _int(row.get("round_id"), 0)
            if w is None or l is None or t is None:
                skip["bad_id"] += 1
                continue
            if w in UNKNOWN_PLAYER_IDS or l in UNKNOWN_PLAYER_IDS:
                skip["unknown_player_id"] += 1
                continue
            pw = players_by_id.get(w)
            pl = players_by_id.get(l)
            if not pw or not pl:
                skip["missing_player_row"] += 1
                continue
            if "/" in (pw.get("name") or "") or "/" in (pl.get("name") or ""):
                skip["doubles_or_team"] += 1
                continue

            tour = tours.get(t)
            if not tour:
                skip["missing_tour"] += 1
                continue
            if not tour.get("is_supported"):
                skip["unsupported_tour"] += 1
                continue

            gd = _to_date(row.get("date"))
            date_ord = gd.toordinal() if gd else tour.get("date_ord")
            if not date_ord:
                skip["missing_event_date"] += 1
                continue
            if date_ord >= max_date_ord:
                continue

            key = (w, l, t, int(rd or 0))
            stat_row = None
            dq = stat_map.get(key)
            if dq and len(dq) > 0:
                stat_row = dq.popleft()

            winner_stats = None
            loser_stats = None
            if stat_row:
                w_w1s, w_w2s, w_fsof, w_w2sof, w_rpw, w_rpwof = stat_row[0], stat_row[1], stat_row[2], stat_row[3], stat_row[4], stat_row[5]
                l_w1s, l_w2s, l_fsof, l_w2sof, l_rpw, l_rpwof = stat_row[6], stat_row[7], stat_row[8], stat_row[9], stat_row[10], stat_row[11]
                w_fs = stat_row[12] if len(stat_row) >= 14 else 0
                l_fs = stat_row[13] if len(stat_row) >= 14 else 0
                winner_stats = (
                    w_w1s + w_w2s,
                    w_fsof + w_w2sof,
                    w_rpw,
                    w_rpwof,
                    w_fs,
                    w_fsof,
                    w_w1s,
                    w_w2sof,
                    w_w2s,
                )
                loser_stats = (
                    l_w1s + l_w2s,
                    l_fsof + l_w2sof,
                    l_rpw,
                    l_rpwof,
                    l_fs,
                    l_fsof,
                    l_w1s,
                    l_w2sof,
                    l_w2s,
                )
                if loser_stats[1] == 0 or loser_stats[3] == 0:
                    loser_stats = None
                if winner_stats[1] == 0 or winner_stats[3] == 0:
                    winner_stats = None

            events.append(
                HistoryEvent(
                    date_ord=int(date_ord),
                    winner_id=int(w),
                    loser_id=int(l),
                    surface=str(tour.get("surface") or "N/A"),
                    tour_key=str(tour.get("tour_key") or ""),
                    tournament_speed_key=str(tour.get("tournament_speed_key") or ""),
                    winner_stats=winner_stats,
                    loser_stats=loser_stats,
                )
            )
    events.sort(key=lambda e: e.date_ord)
    return events, skip


def _log_loss(prob: float) -> float:
    p = _safe_prob(prob)
    return -math.log(p)


def _summarize_calibration(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    bins = [(x / 100.0, (x + 5) / 100.0) for x in range(50, 95, 5)]
    out = []
    for lo, hi in bins:
        sample = [r for r in rows if lo <= r["fav_prob"] < hi]
        if not sample:
            out.append({"bin": f"{int(lo*100)}-{int(hi*100)}%", "n": 0, "pred": None, "actual": None, "gap": None})
            continue
        avg_pred = mean(r["fav_prob"] for r in sample)
        actual = mean(float(r["fav_won"]) for r in sample)
        out.append(
            {
                "bin": f"{int(lo*100)}-{int(hi*100)}%",
                "n": len(sample),
                "pred": avg_pred,
                "actual": actual,
                "gap": actual - avg_pred,
            }
        )
    return out


# Suppress when model and Pinnacle disagree on favourite pricing by >10pp.
# Phantom underdog edges (model 1.15 vs Pin 1.02) cause guaranteed losses.
MISPRICE_IMPLIED_GAP_PP = 0.10
# Skip matches where model favourite odds < threshold. Model cannot price extreme mismatches.
DEFAULT_MISPRICE_MODEL_FAV_ODDS_MIN = 1.25


def _is_mispriced_model_fav_odds(row: dict[str, Any], min_odds: float = DEFAULT_MISPRICE_MODEL_FAV_ODDS_MIN) -> bool:
    """Exclude match if model favourite odds < min_odds (both sides unreliable)."""
    p = _safe_prob(float(row.get("p_model_winner") or 0.5))
    our_odds_w = 1.0 / p
    our_odds_l = 1.0 / _safe_prob(1.0 - p)
    model_fav_odds = min(our_odds_w, our_odds_l)
    return model_fav_odds < min_odds


def _is_mispriced_pinnacle_fav_odds(row: dict[str, Any], min_odds: float = DEFAULT_MISPRICE_MODEL_FAV_ODDS_MIN) -> bool:
    """Exclude match if Pinnacle favourite odds < min_odds. Market massacre = don't bet the dog."""
    psw = float(row.get("psw") or 0)
    psl = float(row.get("psl") or 0)
    if psw <= 0 or psl <= 0:
        return False
    pin_fav_odds = min(psw, psl)
    return pin_fav_odds < min_odds


def _is_mispriced_model_vs_pinnacle(row: dict[str, Any]) -> bool:
    """Exclude match if model fav implied and Pinnacle fav implied differ by >10pp."""
    p_model = _safe_prob(float(row.get("p_model_winner") or 0.5))
    p_pin = _safe_prob(float(row.get("pinnacle_prob_novig") or 0.5))
    model_fav_implied = max(p_model, 1.0 - p_model)
    pin_fav_implied = max(p_pin, 1.0 - p_pin)
    return abs(model_fav_implied - pin_fav_implied) > MISPRICE_IMPLIED_GAP_PP


def _policy_excludes_atp500_hard_short_favorite(
    row: dict[str, Any],
    max_favorite_odds: float,
    allowed_confidence: set[str],
) -> bool:
    if str(row.get("series") or "") != "ATP500":
        return False
    if str(row.get("surface") or "") != "Hard":
        return False
    conf = str(row.get("confidence") or "").strip().lower()
    if conf not in allowed_confidence:
        return False
    p = _safe_prob(float(row.get("p_model_winner") or 0.5))
    fav_odds = 1.0 / max(p, 1.0 - p)
    return fav_odds < max_favorite_odds


def _roi_for_threshold(
    rows: list[dict[str, Any]],
    threshold: float,
    policy_exclude_fn: Callable[[dict[str, Any]], bool] | None = None,
) -> dict[str, Any]:
    bets = 0
    profit = 0.0
    losses = 0
    max_losses = 0
    wins = 0

    for r in rows:
        if policy_exclude_fn and policy_exclude_fn(r):
            continue
        v_w = r["value_w"]
        v_l = r["value_l"]
        if v_w >= v_l:
            v = v_w
            odds = r["psw"]
            won = True
        else:
            v = v_l
            odds = r["psl"]
            won = False
        if v <= threshold:
            continue

        bets += 1
        if won:
            wins += 1
            profit += odds - 1.0
            losses = 0
        else:
            profit -= 1.0
            losses += 1
            if losses > max_losses:
                max_losses = losses

    roi = (profit / bets) if bets > 0 else 0.0
    return {
        "threshold": threshold,
        "bets": bets,
        "wins": wins,
        "profit_units": profit,
        "roi": roi,
        "max_losing_streak": max_losses,
    }


def _segment_stats(
    rows: list[dict[str, Any]],
    threshold: float,
    key: str,
    policy_exclude_fn: Callable[[dict[str, Any]], bool] | None = None,
) -> list[tuple[str, int, float, float]]:
    out = []
    buckets: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        buckets[str(r.get(key) or "N/A")].append(r)
    for bucket, sample in sorted(buckets.items(), key=lambda x: (-len(x[1]), x[0])):
        ll = mean(_log_loss(x["p_model_winner"]) for x in sample) if sample else 0.0
        roi = _roi_for_threshold(sample, threshold, policy_exclude_fn)["roi"]
        out.append((bucket, len(sample), ll, roi))
    return out


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "date",
        "tournament",
        "surface",
        "round",
        "series",
        "player1",
        "player2",
        "player1_id",
        "player2_id",
        "our_prob",
        "our_prob_raw",
        "our_odds",
        "pinnacle_odds",
        "pinnacle_odds_loser",
        "actual_winner",
        "value_pct",
        "bet_side",
        "bet_result",
        "policy_excluded",
        "pinnacle_prob_novig",
        "model_favorite",
        "model_favorite_prob",
        "p_serve_return",
        "p_elo",
        "p_rank",
        "p_a",
        "p_b",
        "confidence",
        "score",
    ]
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k) for k in fields})


def main() -> None:
    parser = argparse.ArgumentParser(description="Backtest fair-odds model vs historical Pinnacle closing odds.")
    parser.add_argument(
        "--files",
        nargs="*",
        default=None,
        help="Backtest XLSX files (default: data/backtest/atp-2024.xlsx data/backtest/atp-2025.xlsx)",
    )
    parser.add_argument("--include-2026", action="store_true", help="Also include data/backtest/atp-2026.xlsx")
    parser.add_argument("--thresholds", default="2,5,10", help="Value%% thresholds (percent, comma-separated)")
    parser.add_argument("--limit-matches", type=int, default=None, help="Process first N matches after sorting (debug)")
    parser.add_argument("--dry-run", action="store_true", help="Do not write CSV/log outputs.")
    parser.add_argument("--v2", action="store_true", help="Use hybrid_v2 SPW-inversion p_a/p_b formula (fixes double-counting).")
    parser.add_argument("--no-matchup", action="store_true", help="Disable matchup model; use legacy/hybrid_v2 p_a/p_b (for A/B comparison).")
    parser.add_argument("--no-decomposed-return", action="store_true", help="Disable return decomposition in matchup model; use aggregate return_pct (for A/B comparison).")
    parser.add_argument("--filter-hard-m1000", action="store_true", help="Restrict ROI/calibration to Hard | Masters 1000 only (live pipeline segment).")
    parser.add_argument("--temperature", type=float, default=None, metavar="T", help="Apply temperature scaling to final probs (e.g. 1.18). Output: backtest-results-YYYY-calibrated.csv")
    parser.add_argument("--show-unmatched", action="store_true", help="Print unmatched tennis-data names.")
    parser.add_argument("--unmatched-limit", type=int, default=40)
    parser.add_argument("--enable-cpi-overlay", action="store_true", help="Enable optional CPI/surface-speed overlay in model probabilities.")
    parser.add_argument(
        "--cpi-file",
        default=str((ROOT / "data" / "backtest" / "tennisabstract-atp-surface-speed.csv").as_posix()),
        help="CPI CSV path (default: data/backtest/tennisabstract-atp-surface-speed.csv).",
    )
    parser.add_argument("--cpi-fallback-years", type=int, default=DEFAULT_CPI_LOOKBACK_YEARS)
    parser.add_argument("--cpi-match-weight", type=float, default=DEFAULT_CPI_MATCH_WEIGHT)
    parser.add_argument("--cpi-match-cap", type=float, default=DEFAULT_CPI_MATCH_CAP)
    parser.add_argument("--cpi-z-cap", type=float, default=DEFAULT_CPI_Z_CAP)
    parser.add_argument("--no-tournament-speed-overlay", action="store_true", help="Disable tournament-speed overlay based on prior editions of the same event.")
    parser.add_argument("--tournament-speed-point-weight", type=float, default=DEFAULT_TOURNAMENT_SPEED_POINT_WEIGHT)
    parser.add_argument(
        "--policy-exclude-atp500-hard-short-favs",
        action="store_true",
        help="Exclude ATP500 Hard short-favorite bets from ROI/segment outputs (policy mode).",
    )
    parser.add_argument(
        "--policy-short-fav-max-odds",
        type=float,
        default=1.8,
        help="Policy: max favorite odds for ATP500 Hard exclusion (default: 1.8).",
    )
    parser.add_argument(
        "--policy-short-fav-confidence",
        default="high",
        help="Policy: confidence levels to exclude (comma-separated, default: high).",
    )
    parser.add_argument(
        "--no-misprice-fav-odds",
        action="store_true",
        help="Disable model and Pinnacle fav odds <1.25 filters (for A/B test).",
    )
    parser.add_argument(
        "--misprice-fav-odds-min",
        type=float,
        default=DEFAULT_MISPRICE_MODEL_FAV_ODDS_MIN,
        help="Exclude matches where model fav odds < this (default: 1.25). Try 1.35 for stricter.",
    )
    parser.add_argument(
        "--challenger",
        action="store_true",
        help="Backtest Challenger matches from oncourt_games + pinnacle-odds CSV (requires data/pinnacle-odds-*.csv).",
    )
    args = parser.parse_args()

    backtest_dir = ROOT / "data" / "backtest"
    data_dir = ROOT / "data"
    if args.challenger:
        xlsx_paths = []
    elif args.files:
        xlsx_paths = [Path(p) if Path(p).is_absolute() else (ROOT / p) for p in args.files]
    else:
        xlsx_paths = [backtest_dir / "atp-2024.xlsx", backtest_dir / "atp-2025.xlsx"]
    if args.include_2026 and not args.challenger:
        p2026 = backtest_dir / "atp-2026.xlsx"
        if p2026 not in xlsx_paths:
            xlsx_paths.append(p2026)

    thresholds = []
    for part in str(args.thresholds).split(","):
        part = part.strip()
        if not part:
            continue
        v = _float(part)
        if v is None:
            continue
        thresholds.append(float(v) / 100.0)
    if not thresholds:
        thresholds = list(DEFAULT_THRESHOLDS)

    policy_exclude_fn: Callable[[dict[str, Any]], bool] | None = None
    policy_allowed_conf: set[str] = set()
    policy_max_odds = max(1.01, float(args.policy_short_fav_max_odds))

    def _combined_policy_exclude(row: dict[str, Any]) -> bool:
        # Always apply misprice filter (model vs Pinnacle fav implied gap >10pp)
        if _is_mispriced_model_vs_pinnacle(row):
            return True
        # Skip matches where model favourite odds < threshold (both sides unreliable)
        # Set --no-misprice-fav-odds to disable for A/B test
        min_odds = getattr(args, "misprice_fav_odds_min", DEFAULT_MISPRICE_MODEL_FAV_ODDS_MIN)
        if not getattr(args, "no_misprice_fav_odds", False) and _is_mispriced_model_fav_odds(row, min_odds):
            return True
        # Skip matches where Pinnacle favourite odds < threshold. Market massacre = don't bet the dog.
        if not getattr(args, "no_misprice_fav_odds", False) and _is_mispriced_pinnacle_fav_odds(row, min_odds):
            return True
        if policy_exclude_fn:
            return policy_exclude_fn(row)
        return False

    if args.policy_exclude_atp500_hard_short_favs:
        policy_allowed_conf = {
            x.strip().lower()
            for x in str(args.policy_short_fav_confidence).split(",")
            if x.strip()
        }
        if not policy_allowed_conf:
            policy_allowed_conf = {"high"}

        def _policy_exclude(row: dict[str, Any]) -> bool:
            return _policy_excludes_atp500_hard_short_favorite(
                row,
                max_favorite_odds=policy_max_odds,
                allowed_confidence=policy_allowed_conf,
            )

        policy_exclude_fn = _policy_exclude

    # Use combined filter (misprice + optional ATP500) for ROI/segments
    effective_policy_exclude_fn = _combined_policy_exclude

    cpi_enabled = bool(args.enable_cpi_overlay)
    cpi_file = Path(args.cpi_file) if Path(args.cpi_file).is_absolute() else (ROOT / args.cpi_file)
    cpi_fallback_years = max(0, int(args.cpi_fallback_years))
    cpi_match_weight = _clamp(float(args.cpi_match_weight), 0.0, 0.15)
    cpi_match_cap = _clamp(float(args.cpi_match_cap), 0.0, 0.08)
    cpi_z_cap = _clamp(float(args.cpi_z_cap), 0.2, 6.0)
    tournament_speed_enabled = not bool(getattr(args, "no_tournament_speed_overlay", False))
    tournament_speed_point_weight = _clamp(float(args.tournament_speed_point_weight), 0.0, 0.40)
    cpi_lookup: dict[tuple[int, str, str], float] = {}
    cpi_year_surface_stats: dict[tuple[int, str], tuple[float, float]] = {}
    cpi_surface_stats: dict[str, tuple[float, float]] = {}

    if args.challenger:
        data_oncourt = ROOT / "data" / "oncourt"
        pinnacle_dir = ROOT / "data"
        print("Loading Challenger matches (oncourt_games + pinnacle-odds CSV)...")
        matches, skip_from_xlsx, short_names = _load_challenger_matches(
            data_oncourt / "games_atp.csv",
            data_oncourt / "tours_atp.csv",
            data_oncourt / "players_atp.csv",
            data_oncourt / "courts.csv",
            pinnacle_dir,
        )
    else:
        print("Loading tennis-data files...")
        matches, skip_from_xlsx, short_names = _load_backtest_matches(xlsx_paths)
    if args.limit_matches:
        matches = matches[: args.limit_matches]
    if not matches:
        print("No usable matches loaded.")
        return

    if cpi_enabled:
        cpi_lookup, cpi_year_surface_stats, cpi_surface_stats = _load_cpi_lookup(cpi_file)
        print(
            "CPI overlay: "
            f"enabled rows={len(cpi_lookup):,} file={cpi_file} "
            f"fallback_years={cpi_fallback_years} weight={cpi_match_weight:.3f} cap={cpi_match_cap:.3f} z_cap={cpi_z_cap:.2f}"
        )
    else:
        print("CPI overlay: disabled")
    print(
        "Tournament speed overlay: "
        f"{'enabled' if tournament_speed_enabled else 'disabled'} "
        f"(weight={tournament_speed_point_weight:.3f}, lookback_days={TOURNAMENT_SPEED_LOOKBACK_DAYS}, "
        f"min_matches={MIN_TOURNAMENT_SPEED_MATCHES})"
    )

    if getattr(args, "v2", False):
        if HAS_HYBRID_V2:
            print("V2 formula: SPW-inversion p_a/p_b (hybrid_v2.compute_point_probs_bc) enabled.")
        else:
            print("V2 formula requested but hybrid_v2 not found; falling back to V4.")

    if args.temperature is not None:
        print(f"Temperature scaling: T={args.temperature:.2f} (output: *-calibrated.csv)")

    print(f"Backtest matches loaded: {len(matches):,}")
    if skip_from_xlsx:
        print(f"  Skipped while reading XLSX: {dict(skip_from_xlsx)}")

    data_oncourt = ROOT / "data" / "oncourt"
    players_path = data_oncourt / "players_atp.csv"
    games_path = data_oncourt / "games_atp.csv"
    stat_path = data_oncourt / "stat_atp.csv"
    tours_path = data_oncourt / "tours_atp.csv"
    courts_path = data_oncourt / "courts.csv"
    categories_path = data_oncourt / "categories_atp.csv"

    for p in [players_path, games_path, stat_path, tours_path, courts_path]:
        if not p.exists():
            raise FileNotFoundError(f"Missing required OnCourt file: {p}")

    print("Loading OnCourt players/tours...")
    oncourt_players = _load_oncourt_players(players_path)
    players_by_id = {int(r["id"]): r for r in oncourt_players}
    birthdate_by_player = {int(r["id"]): r.get("birthdate") for r in oncourt_players}
    lefties = _load_lefties(categories_path, set(players_by_id.keys()))
    tours = _load_tour_info(tours_path, courts_path)
    popularity_by_pid = _load_player_popularity(games_path, players_by_id, tours)

    if args.challenger:
        name_map = {}
        map_report = {"source_names": 0, "mapped_names": 0, "unmapped_names": 0}
        print("Challenger: using OnCourt IDs directly (no tennis-data mapping).")
    else:
        print("Mapping tennis-data names to OnCourt IDs...")
        name_map, map_report = _build_tennis_data_name_map(
            short_names,
            oncourt_players,
            popularity_by_pid=popularity_by_pid,
        )
        print(f"  OnCourt players: {len(oncourt_players):,}")
        print(f"  Lefties from categories_atp.csv: {len(lefties):,}")
        print(f"  OnCourt index players (singles only): {map_report['oncourt_index_players']:,}")
        print(f"  Popularity rows: {len(popularity_by_pid):,}")
        print(f"  Tennis-data short names: {map_report['source_names']:,}")
        print(f"  Mapped short names: {map_report['mapped_names']:,}")
        print(f"  Unmapped short names: {map_report['unmapped_names']:,}")
        print(f"  Mapping methods: {map_report['method_counts']}")
        if args.show_unmatched and map_report.get("unmapped_examples"):
            print("  Unmapped examples:")
            for nm in map_report["unmapped_examples"][: args.unmatched_limit]:
                print(f"    - {nm}")

    skipped = Counter()
    filtered_matches: list[BacktestMatch] = []
    for m in matches:
        if args.challenger:
            wid = m.winner_id
            lid = m.loser_id
        else:
            wid = name_map.get(m.winner_name_short)
            lid = name_map.get(m.loser_name_short)
        if wid is None:
            skipped["unmapped_winner"] += 1
            continue
        if lid is None:
            skipped["unmapped_loser"] += 1
            continue
        if wid == lid:
            skipped["same_player_after_mapping"] += 1
            continue
        if wid in UNKNOWN_PLAYER_IDS or lid in UNKNOWN_PLAYER_IDS:
            skipped["unknown_player_id"] += 1
            continue
        if "/" in (players_by_id.get(wid, {}).get("name") or "") or "/" in (players_by_id.get(lid, {}).get("name") or ""):
            skipped["mapped_to_doubles_team"] += 1
            continue
        m.winner_id = wid
        m.loser_id = lid
        filtered_matches.append(m)

    if not filtered_matches:
        print("No matches left after ID mapping.")
        return

    print(f"Matches after ID mapping: {len(filtered_matches):,}")
    if skipped:
        print(f"  Skipped after mapping: {dict(skipped)}")

    print("Loading historical OnCourt stat/games...")
    stat_map = _load_stat_map(stat_path)
    max_date_ord = max(m.date_ord for m in filtered_matches)
    history_events, history_skip = _history_from_oncourt(games_path, players_by_id, tours, stat_map, max_date_ord=max_date_ord)
    print(f"  History events loaded: {len(history_events):,}")
    if history_skip:
        print(f"  History skipped: {dict(history_skip)}")

    games_by_player: dict[int, list[dict[str, Any]]] | None = None
    tour_surface_map: dict[int, str] | None = None
    if HAS_FATIGUE_MODEL:
        backtest_player_ids = {m.winner_id for m in filtered_matches} | {m.loser_id for m in filtered_matches}
        games_by_player, tour_surface_map = _load_games_for_fatigue(
            games_path, tours, max_date_ord, player_ids=backtest_player_ids
        )
        total_games = sum(len(v) for v in games_by_player.values()) // 2
        print(f"  Games for fatigue: {total_games:,} rows, {len(games_by_player):,} players")

    print("Running strict chronological backtest (no future matches in features)...")
    state = ModelState()
    hist_idx = 0
    results: list[dict[str, Any]] = []
    confidence_counts = Counter()
    by_year_rows: dict[int, list[dict[str, Any]]] = defaultdict(list)
    matchup_method_by_year: dict[int, dict[str, int]] = defaultdict(lambda: {"both_decomposed": 0, "one_decomposed": 0, "neither_decomposed": 0, "legacy": 0})

    for m in filtered_matches:
        while hist_idx < len(history_events) and history_events[hist_idx].date_ord < m.date_ord:
            state.update(history_events[hist_idx], lefties)
            hist_idx += 1

        p_winner, confidence, debug = _compute_match_probability(
            m,
            state,
            birthdate_by_player,
            lefties,
            use_v2_formula=getattr(args, "v2", False),
            use_matchup_model=not getattr(args, "no_matchup", False),
            use_decomposed_return=not getattr(args, "no_decomposed_return", False),
            cpi_enabled=cpi_enabled,
            cpi_lookup=cpi_lookup,
            cpi_year_surface_stats=cpi_year_surface_stats,
            cpi_surface_stats=cpi_surface_stats,
            cpi_fallback_years=cpi_fallback_years,
            cpi_match_weight=cpi_match_weight,
            cpi_match_cap=cpi_match_cap,
            cpi_z_cap=cpi_z_cap,
            tournament_speed_enabled=tournament_speed_enabled,
            tournament_speed_point_weight=tournament_speed_point_weight,
            games_by_player=games_by_player,
            tour_surface_map=tour_surface_map,
        )
        if args.temperature is not None:
            p_winner = _apply_temperature(p_winner, args.temperature)
        confidence_counts[confidence] += 1
        mm = debug.get("matchup_method") or "legacy"
        if mm in matchup_method_by_year[m.year]:
            matchup_method_by_year[m.year][mm] += 1

        our_odds_w = 1.0 / _safe_prob(p_winner)
        our_odds_l = 1.0 / _safe_prob(1.0 - p_winner)
        p_pin_w_raw = 1.0 / m.psw
        p_pin_l_raw = 1.0 / m.psl
        pin_sum = p_pin_w_raw + p_pin_l_raw
        p_pin_w = p_pin_w_raw / pin_sum if pin_sum > 0 else 0.5

        value_w = (m.psw / our_odds_w) - 1.0
        value_l = (m.psl / our_odds_l) - 1.0
        best_side = "winner" if value_w >= value_l else "loser"
        best_value = value_w if value_w >= value_l else value_l
        best_result = "win" if best_side == "winner" else "loss"

        record = {
            "year": m.year,
            "date_ord": m.date_ord,
            "date": m.date_iso,
            "tournament": m.tournament,
            "surface": m.surface,
            "round": m.round_name,
            "series": m.series,
            "player1": players_by_id[m.winner_id]["name"],
            "player2": players_by_id[m.loser_id]["name"],
            "player1_id": m.winner_id,
            "player2_id": m.loser_id,
            "our_prob": round(p_winner, 6),
            "our_prob_raw": round(float(debug.get("p_raw") or p_winner), 6),
            "our_odds": round(our_odds_w, 4),
            "pinnacle_odds": round(m.psw, 4),
            "actual_winner": players_by_id[m.winner_id]["name"],
            "value_pct": round(best_value * 100.0, 3),
            "bet_side": best_side,
            "bet_result": best_result,
            "pinnacle_prob_novig": round(p_pin_w, 6),
            "model_favorite": players_by_id[m.winner_id]["name"] if p_winner >= 0.5 else players_by_id[m.loser_id]["name"],
            "model_favorite_prob": round(max(p_winner, 1.0 - p_winner), 6),
            "p_model_winner": p_winner,
            "p_pin_winner": p_pin_w,
            "fav_prob": max(p_winner, 1.0 - p_winner),
            "fav_won": 1 if p_winner >= 0.5 else 0,
            "psw": m.psw,
            "psl": m.psl,
            "value_w": value_w,
            "value_l": value_l,
            "confidence": confidence,
            "p_serve_return": debug.get("p_serve_return"),
            "p_elo": debug.get("p_elo"),
            "p_rank": debug.get("p_rank"),
            "p_a": debug.get("p_a"),
            "p_b": debug.get("p_b"),
            "p_raw": float(debug.get("p_raw") or p_winner),
            "cpi_value": debug.get("cpi_value"),
            "cpi_year": debug.get("cpi_year"),
            "cpi_z": debug.get("cpi_z"),
            "cpi_delta": debug.get("cpi_delta"),
            "tournament_speed_signal": debug.get("tournament_speed_signal"),
            "tournament_speed_multiplier": debug.get("tournament_speed_multiplier"),
            "tournament_speed_venue_spw": debug.get("tournament_speed_venue_spw"),
            "tournament_speed_match_count": debug.get("tournament_speed_match_count"),
            "score": m.score,
        }
        results.append(record)
        by_year_rows[m.year].append(record)

    if not results:
        print("No model results produced.")
        return

    filter_hard_m1000 = getattr(args, "filter_hard_m1000", False)
    eval_results = results
    if filter_hard_m1000:
        eval_results = [r for r in results if r["surface"] == "Hard" and r["series"] == "Masters 1000"]
        print(f"Filter: Hard | Masters 1000 only ({len(eval_results):,} of {len(results):,} matches)")

    print(f"Completed predictions: {len(results):,}")
    print(f"  Confidence: {dict(confidence_counts)}")
    if matchup_method_by_year:
        print("  Matchup model path by year (decomposed vs fallback):")
        for yr in sorted(matchup_method_by_year.keys()):
            d = matchup_method_by_year[yr]
            total = sum(d.values())
            both = d.get("both_decomposed", 0)
            one = d.get("one_decomposed", 0)
            neither = d.get("neither_decomposed", 0)
            legacy = d.get("legacy", 0)
            pct_both = 100.0 * both / total if total else 0
            pct_fallback = 100.0 * (one + neither + legacy) / total if total else 0
            print(f"    {yr}: both_decomposed={both} ({pct_both:.1f}%), fallback/mixed={one + neither + legacy} ({pct_fallback:.1f}%)")
    if cpi_enabled:
        cpi_resolved = sum(1 for r in results if r.get("cpi_value") is not None)
        cpi_adjusted = sum(1 for r in results if abs(float(r.get("cpi_delta") or 0.0)) > 1e-9)
        print(
            f"  CPI overlay usage: resolved={cpi_resolved}/{len(results)} "
            f"adjusted={cpi_adjusted} (weight={cpi_match_weight:.3f}, cap={cpi_match_cap:.3f})"
        )
    if tournament_speed_enabled:
        speed_resolved = sum(1 for r in results if abs(float(r.get("tournament_speed_signal") or 0.0)) > 1e-9)
        speed_adjusted = sum(1 for r in results if abs(float(r.get("tournament_speed_multiplier") or 0.0)) > 1e-9)
        print(
            f"  Tournament speed usage: resolved={speed_resolved}/{len(results)} "
            f"adjusted={speed_adjusted} (weight={tournament_speed_point_weight:.3f})"
        )
    misprice_min = getattr(args, "misprice_fav_odds_min", DEFAULT_MISPRICE_MODEL_FAV_ODDS_MIN)
    misprice_excluded = sum(1 for r in eval_results if _is_mispriced_model_vs_pinnacle(r))
    misprice_model_fav_excluded = sum(
        1 for r in eval_results if _is_mispriced_model_fav_odds(r, misprice_min)
    )
    misprice_pin_fav_excluded = sum(
        1 for r in eval_results if _is_mispriced_pinnacle_fav_odds(r, misprice_min)
    )
    excluded_count = sum(1 for r in eval_results if effective_policy_exclude_fn(r))
    print(
        f"  Misprice filter (model vs Pin fav implied gap >{MISPRICE_IMPLIED_GAP_PP*100:.0f}pp): excluded={misprice_excluded}"
    )
    print(
        f"  Misprice filter (model fav odds <{misprice_min}): excluded={misprice_model_fav_excluded}"
    )
    print(
        f"  Misprice filter (Pinnacle fav odds <{misprice_min}): excluded={misprice_pin_fav_excluded}"
    )
    if policy_exclude_fn:
        print(
            f"  ATP500 Hard short-fav filter: fav_odds<{policy_max_odds:.2f}, conf={sorted(policy_allowed_conf)}"
        )
    print(f"  Total excluded from ROI: {excluded_count}")

    ll_ours = mean(_log_loss(r["p_model_winner"]) for r in eval_results)
    ll_pin = mean(_log_loss(r["p_pin_winner"]) for r in eval_results)
    brier_ours = mean((1.0 - r["p_model_winner"]) ** 2 for r in eval_results)
    brier_pin = mean((1.0 - r["p_pin_winner"]) ** 2 for r in eval_results)
    acc_ours = mean(1.0 if r["p_model_winner"] >= 0.5 else 0.0 for r in eval_results)
    acc_pin = mean(1.0 if r["p_pin_winner"] >= 0.5 else 0.0 for r in eval_results)

    print("\n=== Model vs Pinnacle (No-Vig) ===")
    print(f"Matches evaluated: {len(eval_results):,}")
    print(f"Log loss: ours={ll_ours:.5f} | pinnacle={ll_pin:.5f} | delta={ll_ours - ll_pin:+.5f}")
    print(f"Brier score: ours={brier_ours:.5f} | pinnacle={brier_pin:.5f} | delta={brier_ours - brier_pin:+.5f}")
    print(f"Accuracy (favorite wins): ours={acc_ours:.3%} | pinnacle={acc_pin:.3%}")

    print("\n=== Calibration (favorite probability bins) ===")
    for row in _summarize_calibration(eval_results):
        if row["n"] == 0:
            print(f"{row['bin']:>8}  n=0")
        else:
            print(
                f"{row['bin']:>8}  n={row['n']:4d}  pred={row['pred']:.3f}  actual={row['actual']:.3f}  gap={row['gap']:+.3f}"
            )

    print("\n=== ROI by Value% threshold ===")
    for th in thresholds:
        r = _roi_for_threshold(eval_results, th, effective_policy_exclude_fn)
        print(
            f"Value>{th*100:.1f}%  bets={r['bets']:4d}  wins={r['wins']:4d}  "
            f"ROI={r['roi']:+.3%}  P/L={r['profit_units']:+.2f}u  "
            f"maxLosingStreak={r['max_losing_streak']}"
        )

    base_threshold = thresholds[1] if len(thresholds) >= 2 else thresholds[0]
    print(f"\n=== Segmentation (threshold {base_threshold*100:.1f}%) by Surface ===")
    for surf, n, ll, roi in _segment_stats(eval_results, base_threshold, "surface", effective_policy_exclude_fn):
        print(f"{surf:>8}  n={n:4d}  logloss={ll:.5f}  ROI={roi:+.3%}")

    print(f"\n=== Segmentation (threshold {base_threshold*100:.1f}%) by Series ===")
    for tier, n, ll, roi in _segment_stats(eval_results, base_threshold, "series", effective_policy_exclude_fn):
        print(f"{tier:>12}  n={n:4d}  logloss={ll:.5f}  ROI={roi:+.3%}")

    print("\n=== Non-Technical Summary ===")
    if ll_ours < ll_pin:
        print("The model beat Pinnacle's closing no-vig line on log loss in this sample.")
    else:
        print("The model did not beat Pinnacle's closing no-vig line on log loss in this sample.")
    best_roi = max((_roi_for_threshold(eval_results, th, effective_policy_exclude_fn)["roi"], th) for th in thresholds)
    print(f"Best flat-stake ROI occurred at Value>{best_roi[1]*100:.1f}% with ROI {best_roi[0]:+.2%}.")
    print("Treat this as a validation signal, not production proof, until you repeat across periods and check stability.")

    if not args.dry_run:
        for year, rows in sorted(by_year_rows.items()):
            out_rows = []
            threshold_default = 0.05
            for r in rows:
                side = "winner" if r["value_w"] >= r["value_l"] else "loser"
                value = r["value_w"] if r["value_w"] >= r["value_l"] else r["value_l"]
                bet_result = "skip"
                policy_excluded = bool(effective_policy_exclude_fn and effective_policy_exclude_fn(r))
                if value > threshold_default and not policy_excluded:
                    bet_result = "win" if side == "winner" else "loss"
                elif value > threshold_default and policy_excluded:
                    bet_result = "policy_skip"
                out_rows.append(
                    {
                        "date": r["date"],
                        "tournament": r["tournament"],
                        "surface": r["surface"],
                        "round": r["round"],
                        "series": r["series"],
                        "player1": r["player1"],
                        "player2": r["player2"],
                        "player1_id": r["player1_id"],
                        "player2_id": r["player2_id"],
                        "our_prob": r["our_prob"],
                        "our_prob_raw": round(r["p_raw"], 6),
                        "our_odds": r["our_odds"],
                        "pinnacle_odds": r["pinnacle_odds"],
                        "pinnacle_odds_loser": round(r["psl"], 4),
                        "actual_winner": r["actual_winner"],
                        "value_pct": round(value * 100.0, 3),
                        "bet_side": side,
                        "bet_result": bet_result,
                        "policy_excluded": policy_excluded,
                        "pinnacle_prob_novig": r["pinnacle_prob_novig"],
                        "model_favorite": r["model_favorite"],
                        "model_favorite_prob": r["model_favorite_prob"],
                        "p_serve_return": round(float(r.get("p_serve_return") or 0.0), 6),
                        "p_elo": round(float(r.get("p_elo") or 0.0), 6),
                        "p_rank": "" if r.get("p_rank") is None else round(float(r["p_rank"]), 6),
                        "p_a": r.get("p_a"),
                        "p_b": r.get("p_b"),
                        "confidence": r.get("confidence"),
                        "score": r.get("score") or "",
                    }
                )
            suffix = "-v2" if getattr(args, "v2", False) else ""
            if args.temperature is not None:
                suffix = suffix + "-calibrated" if suffix else "-calibrated"
            prefix = "challenger-" if args.challenger else ""
            out_path = backtest_dir / f"backtest-results-{prefix}{year}{suffix}.csv"
            _write_csv(out_path, out_rows)
            print(f"Wrote {len(out_rows):,} rows -> {out_path}")

        log_path = backtest_dir / ("backtest-log-challenger.txt" if args.challenger else "backtest-log.txt")
        with open(log_path, "w", encoding="utf-8") as f:
            f.write("Backtest run log\n")
            f.write(f"Mode: {'Challenger (oncourt + pinnacle-odds)' if args.challenger else 'ATP (tennis-data XLSX)'}\n")
            if xlsx_paths:
                f.write(f"Files: {', '.join(str(p) for p in xlsx_paths)}\n")
            f.write(f"Loaded matches: {len(matches)}\n")
            f.write(f"Mapped matches: {len(filtered_matches)}\n")
            f.write(f"Predicted matches: {len(results)}\n")
            f.write(f"Skip from xlsx: {dict(skip_from_xlsx)}\n")
            f.write(f"Skip after mapping: {dict(skipped)}\n")
            if effective_policy_exclude_fn:
                f.write(
                    "Policy filter: ATP500 Hard short favorites excluded "
                    f"(fav_odds<{policy_max_odds:.2f}, confidence={sorted(policy_allowed_conf)})\n"
                )
            f.write(f"Name mapping report: {map_report}\n")
            if map_report.get("unmapped_examples"):
                f.write("Unmapped short-name examples:\n")
                for nm in map_report["unmapped_examples"][:300]:
                    f.write(f"  - {nm}\n")
        print(f"Wrote log -> {log_path}")
    else:
        print("\nDry run: skipped CSV/log file writes.")


if __name__ == "__main__":
    main()
